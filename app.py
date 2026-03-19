#!/usr/bin/env python3
"""Modern Streamlit dashboard for Memo Automator."""

from __future__ import annotations

from datetime import datetime
import glob
import os
from pathlib import Path
import time
import uuid

import streamlit as st

from app_helpers import should_disable_fire_button, verify_password
from app_services import (
    accept_invitation,
    add_user,
    consume_credit,
    create_invitation,
    delete_job,
    delete_user,
    enqueue_job,
    ensure_users_seeded,
    get_db_conn,
    get_invitation,
    get_invitations,
    get_job,
    get_job_queue,
    get_job_staging_dir,
    get_platform_health,
    get_profiles,
    get_recent_runs,
    get_run_artifact_paths,
    get_run_storage_dir,
    get_run_details,
    get_user_credits,
    get_users,
    record_run,
    reset_user_credits,
    save_profile,
    start_background_worker,
    update_job_status,
    update_run_approval,
    update_user,
)
from memo_chef.models import CompUrl, RunRequest, StageUpdate
from memo_chef.pipeline import run_memo_pipeline
from memo_chef.theme import APP_SUBTITLE, APP_TITLE, app_css, info_card, render_hero

# ---------------------------------------------------------------------------
# Animated shrimp chef GIF shown during pipeline runs
# ---------------------------------------------------------------------------
_CHEF_CAPTIONS = [
    "The shrimp is vibing while your memo cooks...",
    "Chef Shrimp is plating your metrics...",
    "Shrimp doing the happy dance for your data...",
    "Shrimply vibing while the numbers simmer...",
    "Party shrimp is prepping the garnish...",
]

_COOKING_SHRIMP_B64: str | None = None


def _load_cooking_shrimp_b64() -> str:
    """Load the cooking shrimp GIF as a base64 data URI (cached)."""
    global _COOKING_SHRIMP_B64
    if _COOKING_SHRIMP_B64 is None:
        import base64
        gif_path = Path(__file__).parent / "assets" / "cooking_shrimp.gif"
        _COOKING_SHRIMP_B64 = base64.b64encode(gif_path.read_bytes()).decode()
    return _COOKING_SHRIMP_B64


def _chef_gif_html() -> str:
    """Return HTML for the animated cooking shrimp chef GIF."""
    import random
    caption = random.choice(_CHEF_CAPTIONS)
    b64 = _load_cooking_shrimp_b64()
    return f"""
<div style="text-align:center; margin: 15px 0;">
  <img src="data:image/gif;base64,{b64}" alt="Chef Shrimp"
       style="max-height:220px; border-radius:12px;
              box-shadow: 0 4px 15px rgba(0,0,0,0.15);" />
  <p style="margin-top:8px; font-size:14px; color:#888;
            font-style:italic;">{caption}</p>
</div>
"""

st.set_page_config(page_title=APP_TITLE, page_icon="✨", layout="wide")
st.markdown(app_css(), unsafe_allow_html=True)

try:
    ensure_users_seeded()
except Exception as e:
    import logging as _logging
    _logging.getLogger(__name__).warning("Failed to seed users from database: %s", e)

# Start background worker for auto-processing queued jobs
start_background_worker()

# --- Invitation sign-up page (no auth required) ---
invite_token = st.query_params.get("invite")
if invite_token:
    render_hero()
    invite = get_invitation(invite_token)
    if invite is None:
        st.error("Invalid invitation link.")
        st.stop()
    if invite["status"] == "accepted":
        st.info("This invitation has already been used. Please sign in.")
        st.stop()
    if invite["status"] == "expired" or invite["expires_at"] < datetime.now(invite["expires_at"].tzinfo):
        st.warning("This invitation has expired. Please contact your administrator for a new one.")
        st.stop()

    st.markdown(
        info_card("Create your account", f"You've been invited as a **{invite['role']}** with **{invite['credits_per_week']}** weekly runs."),
        unsafe_allow_html=True,
    )
    with st.form("invite_signup_form"):
        signup_username = st.text_input("Choose a username")
        signup_password = st.text_input("Password", type="password")
        signup_confirm = st.text_input("Confirm password", type="password")
        signup_submitted = st.form_submit_button("Create account", type="primary")
    if signup_submitted:
        if not signup_username.strip():
            st.error("Username is required.")
        elif len(signup_password) < 6:
            st.error("Password must be at least 6 characters.")
        elif signup_password != signup_confirm:
            st.error("Passwords do not match.")
        else:
            if accept_invitation(invite_token, signup_username.strip(), signup_password):
                st.success("Account created! You can now sign in.")
                st.query_params.clear()
                time.sleep(2)
                st.rerun()
            else:
                st.error("Could not create account. The username may already be taken or the invitation has expired.")
    st.stop()


_CONFIGS_DIR = os.path.join(os.path.dirname(__file__), "configs")


def _list_config_profiles() -> list[str]:
    """Return stem names of YAML files in configs/, sorted."""
    if not os.path.isdir(_CONFIGS_DIR):
        return []
    return sorted(
        os.path.splitext(os.path.basename(p))[0]
        for p in glob.glob(os.path.join(_CONFIGS_DIR, "*.yaml"))
    )


def _config_override_path(profile_name: str | None) -> str | None:
    if not profile_name:
        return None
    path = os.path.join(_CONFIGS_DIR, f"{profile_name}.yaml")
    return path if os.path.exists(path) else None


def _get_api_key() -> str | None:
    try:
        return st.secrets["ANTHROPIC_API_KEY"]
    except (KeyError, FileNotFoundError):
        return None


def _clear_run_state() -> None:
    for key in [
        "memo_bytes",
        "log_bytes",
        "manifest_bytes",
        "filename",
        "n_changes",
        "n_rejected",
        "n_missed",
        "log_lines",
        "warnings",
        "manifest",
        "changes",
    ]:
        st.session_state.pop(key, None)


def _queue_item_from_inputs(
    *,
    memo_file,
    proforma_file,
    schedule_file,
    market_data_file,
    supplemental_file=None,
    supplemental_url: str = "",
    supplemental_brief: str = "",
    comp_urls: list[dict] | None = None,
    auto_generate_comp_slide: bool = False,
    comp_csv_file=None,
    property_name: str,
    property_rename_to: str = "",
    dry_run: bool,
    skip_validation: bool,
    profile_name: str | None,
    config_profile_name: str | None = None,
    use_batch_api: bool = False,
    source_directives: list[dict] | None = None,
) -> dict:
    # Determine supplemental source type
    supp_name = None
    supp_bytes = None
    supp_type = None
    if supplemental_file:
        supp_name = supplemental_file.name
        supp_bytes = supplemental_file.getvalue()
        ext = Path(supp_name).suffix.lower()
        supp_type = {".pdf": "pdf", ".xlsx": "excel", ".xlsm": "excel", ".csv": "csv", ".txt": "text"}.get(ext, "excel")
    elif supplemental_url:
        supp_name = supplemental_url
        supp_type = "url"

    job_id = uuid.uuid4().hex
    staging = get_job_staging_dir(job_id)

    memo_path = str(staging / memo_file.name)
    with open(memo_path, "wb") as f:
        f.write(memo_file.getvalue())

    proforma_path = str(staging / proforma_file.name)
    with open(proforma_path, "wb") as f:
        f.write(proforma_file.getvalue())

    schedule_path = None
    if schedule_file:
        schedule_path = str(staging / schedule_file.name)
        with open(schedule_path, "wb") as f:
            f.write(schedule_file.getvalue())

    market_data_path = None
    if market_data_file:
        market_data_path = str(staging / market_data_file.name)
        with open(market_data_path, "wb") as f:
            f.write(market_data_file.getvalue())

    supp_path = None
    if supp_bytes:
        supp_path = str(staging / supp_name)
        with open(supp_path, "wb") as f:
            f.write(supp_bytes)

    comp_csv_path = None
    if comp_csv_file:
        comp_csv_path = str(staging / comp_csv_file.name)
        with open(comp_csv_path, "wb") as f:
            f.write(comp_csv_file.getvalue())

    return {
        "job_id": job_id,
        "memo_name": memo_file.name,
        "memo_path": memo_path,
        "proforma_name": proforma_file.name,
        "proforma_path": proforma_path,
        "schedule_name": schedule_file.name if schedule_file else None,
        "schedule_path": schedule_path,
        "market_data_name": market_data_file.name if market_data_file else None,
        "market_data_path": market_data_path,
        "supplemental_name": supp_name,
        "supplemental_path": supp_path,
        "supplemental_type": supp_type,
        "supplemental_brief": supplemental_brief or None,
        "property_name": property_name or None,
        "property_rename_to": property_rename_to or None,
        "comp_urls": comp_urls or [],
        "auto_generate_comp_slide": auto_generate_comp_slide,
        "comp_csv_path": comp_csv_path,
        "dry_run": dry_run,
        "skip_validation": skip_validation,
        "use_batch_api": use_batch_api,
        "profile_name": profile_name or "",
        "config_profile_name": config_profile_name or "",
        "source_directives": source_directives or [],
    }


def _persist_result(result, filename: str) -> None:
    st.session_state["memo_bytes"] = result.memo_bytes
    st.session_state["log_bytes"] = result.log_bytes
    st.session_state["manifest_bytes"] = result.manifest_bytes
    st.session_state["filename"] = filename
    st.session_state["n_changes"] = len(result.changes)
    st.session_state["n_rejected"] = len(result.rejected)
    st.session_state["n_missed"] = len(result.missed)
    st.session_state["unvalidated_pages"] = result.unvalidated_pages
    st.session_state["log_lines"] = result.log_lines
    st.session_state["warnings"] = [warning.model_dump() for warning in result.manifest.warnings]
    st.session_state["manifest"] = result.manifest.model_dump()
    st.session_state["changes"] = result.changes


def _execute_job(
    *,
    job: dict,
    username: str,
    credits_per_week: int,
    queue_position: int | None = None,
    queue_total: int | None = None,
) -> bool:
    _clear_run_state()
    api_key = _get_api_key()
    if not api_key:
        st.error("`ANTHROPIC_API_KEY` is not configured in Streamlit secrets.")
        return False

    run_id = uuid.uuid4().hex
    if job.get("job_id"):
        update_job_status(job["job_id"], "running", run_id=run_id)
    started = time.time()
    prefix = ""
    if queue_position is not None and queue_total is not None:
        prefix = f"Queue item {queue_position}/{queue_total} · "
    progress_bar = st.progress(0, text=f"{prefix}Initializing run")
    shrimp_placeholder = st.empty()
    shrimp_placeholder.markdown(_chef_gif_html(), unsafe_allow_html=True)
    status_box = st.empty()
    stage_log = st.empty()
    stage_lines: list[str] = []

    def on_stage(update: StageUpdate) -> None:
        progress_bar.progress(update.percent, text=f"{prefix}{update.label}")
        message = update.detail or update.label
        stage_lines.append(f"{update.percent:>3}%  {message}")
        status_box.caption(f"{prefix}{update.label}")
        stage_log.code("\n".join(stage_lines[-10:]), language=None)

    run_dir = get_run_storage_dir(run_id)

    # Support both new-style (file paths on disk) and old-style (raw bytes) payloads
    if job.get("memo_path") and os.path.isfile(job["memo_path"]):
        memo_path = job["memo_path"]
    else:
        memo_path = str(run_dir / f"input_memo{os.path.splitext(job['memo_name'])[1]}")
        with open(memo_path, "wb") as handle:
            handle.write(job["memo_bytes"])

    if job.get("proforma_path") and os.path.isfile(job["proforma_path"]):
        proforma_path = job["proforma_path"]
    else:
        proforma_path = str(run_dir / f"input_proforma{os.path.splitext(job['proforma_name'])[1]}")
        with open(proforma_path, "wb") as handle:
            handle.write(job["proforma_bytes"])

    schedule_path = None
    if job.get("schedule_path") and os.path.isfile(job["schedule_path"]):
        schedule_path = job["schedule_path"]
    elif job.get("schedule_bytes"):
        schedule_path = str(run_dir / f"input_schedule{os.path.splitext(job['schedule_name'])[1]}")
        with open(schedule_path, "wb") as handle:
            handle.write(job["schedule_bytes"])

    market_data_path = None
    if job.get("market_data_path") and os.path.isfile(job["market_data_path"]):
        market_data_path = job["market_data_path"]
    elif job.get("market_data_bytes"):
        market_data_path = str(run_dir / f"input_market_data{os.path.splitext(job['market_data_name'])[1]}")
        with open(market_data_path, "wb") as handle:
            handle.write(job["market_data_bytes"])

    supplemental_path = None
    supplemental_type = job.get("supplemental_type")
    if supplemental_type == "url":
        supplemental_path = job.get("supplemental_name")  # URL string
    elif job.get("supplemental_path") and os.path.isfile(job["supplemental_path"]):
        supplemental_path = job["supplemental_path"]
    elif job.get("supplemental_bytes"):
        ext = os.path.splitext(job["supplemental_name"])[1] if job.get("supplemental_name") else ".pdf"
        supplemental_path = str(run_dir / f"input_supplemental{ext}")
        with open(supplemental_path, "wb") as handle:
            handle.write(job["supplemental_bytes"])

    comp_url_objects = [CompUrl(**cu) for cu in job.get("comp_urls", [])]

    comp_csv_path = None
    if job.get("comp_csv_path") and os.path.isfile(job["comp_csv_path"]):
        comp_csv_path = job["comp_csv_path"]

    # Build source directives from job payload
    from memo_chef.models import SourceDirective

    source_directives = []
    for sd_dict in job.get("source_directives", []):
        if sd_dict.get("directive", "").strip():
            source_directives.append(SourceDirective(**sd_dict))

    request = RunRequest(
        memo_path=memo_path,
        proforma_path=proforma_path,
        schedule_path=schedule_path,
        market_data_path=market_data_path,
        supplemental_path=supplemental_path,
        supplemental_type=supplemental_type,
        supplemental_brief=job.get("supplemental_brief"),
        comp_urls=comp_url_objects,
        source_directives=source_directives,
        auto_generate_comp_slide=job.get("auto_generate_comp_slide", False),
        comp_csv_path=comp_csv_path,
        output_dir=str(run_dir),
        api_key=api_key,
        config_path=os.path.join(os.path.dirname(__file__), "config.yaml"),
        config_override_path=_config_override_path(job.get("config_profile_name")),
        run_id=run_id,
        property_name=job.get("property_name"),
        property_rename_to=job.get("property_rename_to"),
        dry_run=job.get("dry_run", False),
        skip_validation=job.get("skip_validation", False),
        use_batch_api=job.get("use_batch_api", False),
    )

    try:
        result = run_memo_pipeline(request, callback=on_stage)
        duration = round(time.time() - started, 2)
        progress_bar.progress(100, text=f"{prefix}Run complete")
        shrimp_placeholder.empty()
        status_box.success(f"{prefix}Draft generated successfully.")
        _persist_result(result, job["memo_name"])
        (run_dir / f"memo{os.path.splitext(job['memo_name'])[1]}").write_bytes(result.memo_bytes)
        (run_dir / "change_log.md").write_bytes(result.log_bytes)
        (run_dir / "run_manifest.json").write_bytes(result.manifest_bytes)
        counts = result.manifest.counts
        accuracy = result.manifest.accuracy or {}
        record_run(
            run_id=run_id,
            username=username,
            status=result.manifest.status,
            memo_name=result.manifest.memo_name,
            proforma_name=result.manifest.proforma_name,
            property_name=result.manifest.property_name,
            dry_run=job.get("dry_run", False),
            skip_validation=job.get("skip_validation", False),
            change_count=len(result.changes),
            rejected_count=len(result.rejected),
            missed_count=len(result.missed),
            duration_seconds=duration,
            warnings=st.session_state["warnings"],
            input_tokens=counts.get("input_tokens", 0),
            output_tokens=counts.get("output_tokens", 0),
            estimated_cost_microdollars=counts.get("estimated_cost_microdollars", 0),
            slides_inserted=counts.get("slides_inserted", 0),
            confidence_score=accuracy.get("confidence_score"),
            coverage_pct=accuracy.get("coverage_pct"),
            correction_rate_pct=accuracy.get("correction_rate_pct"),
            run_manifest_json=result.manifest_bytes.decode("utf-8") if result.manifest_bytes else None,
            change_log_html=result.log_bytes.decode("utf-8") if result.log_bytes else None,
        )
        if job.get("job_id"):
            update_job_status(job["job_id"], "completed", run_id=run_id)
        try:
            charged = consume_credit(username, credits_per_week, run_id=run_id)
            if not charged:
                st.warning("The run completed, but weekly credits were already exhausted.")
        except Exception as err:
            st.warning(f"Run completed, but credits could not be updated: {err}")
        return True
    except Exception as err:
        duration = round(time.time() - started, 2)
        progress_bar.progress(100, text=f"{prefix}Run failed")
        shrimp_placeholder.empty()
        status_box.error(f"{prefix}Run failed")
        stage_log.code("\n".join(stage_lines[-10:]), language=None)
        try:
            record_run(
                run_id=run_id,
                username=username,
                status="failed",
                memo_name=job["memo_name"],
                proforma_name=job["proforma_name"],
                property_name=job.get("property_name"),
                dry_run=job.get("dry_run", False),
                skip_validation=job.get("skip_validation", False),
                change_count=0,
                rejected_count=0,
                missed_count=0,
                duration_seconds=duration,
                warnings=[{"stage": "pipeline", "message": str(err)}],
            )
        except Exception:
            pass
        if job.get("job_id"):
            update_job_status(job["job_id"], "failed", run_id=run_id, error_message=str(err))
        st.error(f"{prefix}Run failed: {err}")
        return False


def check_password() -> bool:
    if st.session_state.get("authenticated"):
        return True

    users = get_users()
    if not users:
        st.error("No users configured. Add users to Streamlit secrets or the database.")
        st.stop()

    render_hero()
    st.markdown(
        info_card(
            "Secure workspace",
            "Sign in to access governed memo runs, queue execution, approvals, and operational controls.",
        ),
        unsafe_allow_html=True,
    )

    with st.form("login_form"):
        cols = st.columns([1, 1, 1])
        username = cols[0].text_input("Username")
        password = cols[1].text_input("Password", type="password")
        cols[2].markdown("<div style='height: 1.8rem'></div>", unsafe_allow_html=True)
        submitted = cols[2].form_submit_button("Sign in", type="primary", width="stretch")

    if submitted:
        user_cfg = users.get(username)
        if user_cfg and verify_password(password, user_cfg["password_hash"]):
            st.session_state["authenticated"] = True
            st.session_state["username"] = username
            st.session_state["role"] = user_cfg.get("role", "user")
            st.session_state["credits_per_week"] = int(user_cfg.get("credits_per_week", 5))
            st.rerun()
        st.error("Invalid username or password.")
    st.stop()


if not check_password():
    st.stop()

username = st.session_state["username"]
role = st.session_state["role"]
credits_per_week = st.session_state["credits_per_week"]
credits_error = None

try:
    used, remaining = get_user_credits(username, credits_per_week)
except Exception as err:
    used, remaining = 0, 0
    credits_error = str(err)

with st.sidebar:
    st.markdown(f"### {username}")
    st.caption(f"Role: `{role}`")
    if credits_error:
        st.warning("Credits service unavailable.")
        if st.button("Reconnect services"):
            get_db_conn.clear()
            st.rerun()
    else:
        st.caption(f"{remaining} of {credits_per_week} weekly runs remaining")
        st.progress(
            min(used / credits_per_week, 1.0) if credits_per_week > 0 else 1.0,
            text=f"{used} used this week",
        )
    try:
        queue_count = len(
            [job for job in get_job_queue(None if role == "admin" else username) if job["status"] == "queued"]
        )
    except Exception:
        queue_count = 0
    st.caption(f"Batch queue: {queue_count} item(s)")
    st.divider()
    st.caption("Platform")
    st.write("Reviewable automation, typed configuration, queueing, and traceable outputs.")
    if st.button("Sign out", width="stretch"):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.rerun()

render_hero()

card_cols = st.columns(4)
card_cols[0].markdown(
    info_card("Guardrails", "Two-pass mapping and validation with checkpointed artifacts."),
    unsafe_allow_html=True,
)
card_cols[1].markdown(
    info_card("Operations", "Queue multiple runs and review outcomes from a single console."),
    unsafe_allow_html=True,
)
card_cols[2].markdown(
    info_card("Governance", "Track approval status, reviewer, and warnings per run."),
    unsafe_allow_html=True,
)
card_cols[3].markdown(
    info_card("Brand system", "Refreshed dark UI, cleaner actions, and a premium visual hierarchy."),
    unsafe_allow_html=True,
)

tab_labels = ["New Run", "Run History", "Operations", "How To"] + (["Admin"] if role == "admin" else [])
tabs = st.tabs(tab_labels)


def render_new_run_tab() -> None:
    st.subheader("New run")
    st.caption(APP_SUBTITLE)

    profiles = get_profiles(None if role == "admin" else username)
    profile_lookup = {row["Profile"]: row for row in profiles}
    selected_profile = st.selectbox(
        "Saved profile",
        options=[""] + sorted(profile_lookup.keys()),
        format_func=lambda value: "None" if value == "" else value,
        help="Load saved preferences for property naming and QA behavior.",
    )
    profile = profile_lookup.get(selected_profile, {})

    upload_cols = st.columns(4)
    memo_file = upload_cols[0].file_uploader("Memo deck", type=["pptx"], key="memo_upload")
    proforma_file = upload_cols[1].file_uploader("Proforma", type=["xlsx", "xlsm"], key="proforma_upload")
    schedule_file = upload_cols[2].file_uploader("Schedule (Beta)", type=["mpp"], key="schedule_upload")
    market_data_file = upload_cols[3].file_uploader("Market data (Beta)", type=["xlsx", "xlsm"], key="market_upload")

    # Per-source directives — tell Claude how to use each source
    with st.expander("Directions for Claude (per source)", expanded=False):
        st.caption(
            "Give Claude specific instructions for each source. "
            "E.g., 'Only update revenue section' or 'Ignore occupancy data'."
        )
        directive_cols = st.columns(2)
        proforma_directive = directive_cols[0].text_area(
            "Proforma directions",
            key="proforma_directive",
            placeholder="e.g., Only use Unit Mix and Cash Flow tabs",
            height=68,
        )
        schedule_directive = directive_cols[1].text_area(
            "Schedule directions",
            key="schedule_directive",
            placeholder="e.g., Focus on construction milestones only",
            height=68,
        )
        directive_cols2 = st.columns(2)
        market_data_directive = directive_cols2[0].text_area(
            "Market data directions",
            key="market_data_directive",
            placeholder="e.g., Use for rent trend charts only",
            height=68,
        )
        supplemental_directive = directive_cols2[1].text_area(
            "Supplemental data directions",
            key="supplemental_directive",
            placeholder="e.g., Generate a market summary slide from this data",
            height=68,
        )

    # Supplemental data for slide insertion
    supp_cols = st.columns([2, 2, 3])
    supplemental_file = supp_cols[0].file_uploader(
        "Supplemental data",
        type=["pdf", "xlsx", "xlsm", "csv"],
        key="supplemental_upload",
        help="Upload additional data to generate a new slide (PDF, Excel, or CSV)",
    )
    supplemental_url = supp_cols[1].text_input(
        "Or paste a URL",
        key="supplemental_url",
        placeholder="https://...",
    )
    supplemental_brief = supp_cols[2].text_area(
        "Brief (optional)",
        key="supplemental_brief",
        placeholder="e.g., Show student affluence trends for this market",
        height=80,
    )

    # Comp property URLs for rent scraping
    with st.expander("Comp property URLs (rent scraping)"):
        comp_url_count = st.number_input(
            "Number of comp URLs", min_value=0, max_value=10, value=0, key="comp_url_count",
        )
        comp_url_inputs: list[dict] = []
        for i in range(int(comp_url_count)):
            cols = st.columns([3, 2, 3])
            cu_url = cols[0].text_input(f"URL #{i + 1}", key=f"comp_url_{i}", placeholder="https://...")
            cu_label = cols[1].text_input(f"Label #{i + 1}", key=f"comp_label_{i}", placeholder="e.g. Hub Lexington")
            cu_guidance = cols[2].text_input(
                f"Guidance #{i + 1}", key=f"comp_guidance_{i}", placeholder="e.g. Grab 1BR and 4BR rates",
            )
            if cu_url.strip():
                comp_url_inputs.append({"url": cu_url.strip(), "label": cu_label.strip(), "guidance": cu_guidance.strip()})

    st.markdown("**Comp Slide Builder**")
    auto_comp = st.checkbox("Auto-generate comp slide", value=False, key="auto_comp")
    comp_csv = None
    if auto_comp:
        comp_csv = st.file_uploader("Comp data (CSV)", type=["csv"], key="comp_csv")

    rename_cols = st.columns(2)
    property_name = rename_cols[0].text_input(
        "Property name (as it appears in memo)",
        value=profile.get("Property", ""),
        placeholder="e.g. VERVE Lexington",
        help="The property name currently used in the memo deck. Used for targeting updates.",
    )
    property_rename_to = rename_cols[1].text_input(
        "Rename to (if different in proforma)",
        value=profile.get("Rename To", ""),
        placeholder="e.g. VERVE Pittsburgh",
        help="If the proforma uses a different property name, enter it here. "
             "All occurrences will be renamed before the AI pass.",
    )

    # Smart defaults: auto-detect config profile from proforma tabs
    _auto_profile = ""
    _auto_property = ""
    if proforma_file and not profile.get("Property"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(proforma_file, read_only=True, data_only=True)
            sheet_names = [s.lower() for s in wb.sheetnames]
            proforma_file.seek(0)  # reset for later use
            if any("unit mix" in s or "rent roll" in s for s in sheet_names):
                _auto_profile = "multifamily"
            elif any("senior" in s or "assisted" in s for s in sheet_names):
                _auto_profile = "senior_housing"
            elif any("retail" in s or "office" in s for s in sheet_names):
                _auto_profile = "mixed_use"
            # Try to extract property name from first sheet header
            first_sheet = wb[wb.sheetnames[0]]
            for row in first_sheet.iter_rows(max_row=5, max_col=5, values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and len(cell) > 3 and len(cell) < 60:
                        _auto_property = cell.strip()
                        break
                if _auto_property:
                    break
            wb.close()
        except Exception:
            pass  # silently fail — these are optional hints

    config_profiles = _list_config_profiles()
    # Determine default index: saved profile > auto-detected > none
    _default_profile_idx = 0
    if profile.get("Config Profile") and profile["Config Profile"] in config_profiles:
        _default_profile_idx = config_profiles.index(profile["Config Profile"]) + 1
    elif _auto_profile and _auto_profile in config_profiles:
        _default_profile_idx = config_profiles.index(_auto_profile) + 1
    config_profile_name = st.selectbox(
        "Config profile",
        options=[""] + config_profiles,
        index=_default_profile_idx,
        format_func=lambda v: "Default (config.yaml)" if v == "" else v.replace("_", " ").title(),
        help="Override proforma tabs, model, or other settings for this property type."
             + (f" Auto-detected: {_auto_profile}" if _auto_profile else ""),
    )

    pref_cols = st.columns(3)
    with pref_cols[0]:
        dry_run = st.checkbox(
            "Preview only",
            value=bool(profile.get("Preview Only", False)),
            help="Runs the pipeline without saving final deck changes.",
        )
    with pref_cols[1]:
        skip_validation = st.checkbox(
            "Skip AI validation",
            value=bool(profile.get("Skip QA", False)),
            help="Faster, but less safe. Recommended only for trusted dry runs.",
        )
    with pref_cols[2]:
        use_batch_api = st.checkbox(
            "Batch mode (50% off)",
            value=False,
            help="Submits all mapping chunks to the Anthropic Batch API at 50% cost. "
                 "Results typically return within minutes but may take up to 1 hour.",
        )

    review_cols = st.columns(4)
    review_cols[0].caption("Required inputs: memo + proforma")
    review_cols[1].caption("Optional enrichments: schedule + market data")
    review_cols[2].caption("Artifacts are downloadable after each run")
    review_cols[3].caption("Queue jobs to run sequentially with shared settings")

    save_profile_name = st.text_input(
        "Save current preferences as profile",
        placeholder="e.g. Standard IC review",
    )
    profile_notes = st.text_area(
        "Profile notes",
        placeholder="Optional guidance for this profile",
        height=80,
    )
    profile_cols = st.columns([1, 3])
    if profile_cols[0].button("Save profile", width="stretch"):
        if not save_profile_name.strip():
            st.error("Enter a profile name before saving.")
        else:
            save_profile(
                save_profile_name.strip(),
                username,
                property_name or None,
                dry_run,
                skip_validation,
                profile_notes or None,
                config_profile=config_profile_name or None,
                property_rename_to=property_rename_to or None,
            )
            st.success(f"Saved profile `{save_profile_name.strip()}`.")
            st.rerun()

    action_disabled = should_disable_fire_button(memo_file, proforma_file, remaining, credits_error)
    action_cols = st.columns(2)
    # Collect per-source directives from UI state
    _ui_directives = []
    if proforma_directive.strip():
        _ui_directives.append({
            "source_id": "proforma", "source_type": "proforma_tab",
            "directive": proforma_directive.strip(), "scope": "both",
        })
    if schedule_directive.strip():
        _ui_directives.append({
            "source_id": "schedule", "source_type": "schedule",
            "directive": schedule_directive.strip(), "scope": "both",
        })
    if market_data_directive.strip():
        _ui_directives.append({
            "source_id": "market_data", "source_type": "market_data",
            "directive": market_data_directive.strip(), "scope": "both",
        })
    if supplemental_directive.strip():
        _ui_directives.append({
            "source_id": "supplemental", "source_type": "supplemental",
            "directive": supplemental_directive.strip(), "scope": "both",
        })
    # Comp URL guidance is already captured in comp_url_inputs — add as directives too
    for cu in comp_url_inputs:
        if cu.get("guidance", "").strip():
            _ui_directives.append({
                "source_id": f"comp:{cu.get('label') or cu['url'][:30]}",
                "source_type": "comp_url",
                "directive": cu["guidance"],
                "scope": "both",
            })

    if action_cols[0].button(
        f"Generate draft ({remaining} credits left)" if remaining > 0 else "No credits remaining",
        type="primary",
        disabled=action_disabled,
        width="stretch",
    ):
        job = _queue_item_from_inputs(
            memo_file=memo_file,
            proforma_file=proforma_file,
            schedule_file=schedule_file,
            market_data_file=market_data_file,
            supplemental_file=supplemental_file,
            supplemental_url=supplemental_url,
            supplemental_brief=supplemental_brief,
            comp_urls=comp_url_inputs,
            auto_generate_comp_slide=auto_comp,
            comp_csv_file=comp_csv,
            property_name=property_name,
            property_rename_to=property_rename_to,
            dry_run=dry_run,
            skip_validation=skip_validation,
            profile_name=selected_profile or save_profile_name.strip() or None,
            config_profile_name=config_profile_name or None,
            use_batch_api=use_batch_api,
            source_directives=_ui_directives,
        )
        _execute_job(job=job, username=username, credits_per_week=credits_per_week)

    if action_cols[1].button(
        "Add to queue",
        disabled=action_disabled,
        width="stretch",
    ):
        job = _queue_item_from_inputs(
            memo_file=memo_file,
            proforma_file=proforma_file,
            schedule_file=schedule_file,
            market_data_file=market_data_file,
            supplemental_file=supplemental_file,
            supplemental_url=supplemental_url,
            supplemental_brief=supplemental_brief,
            comp_urls=comp_url_inputs,
            auto_generate_comp_slide=auto_comp,
            comp_csv_file=comp_csv,
            property_name=property_name,
            property_rename_to=property_rename_to,
            dry_run=dry_run,
            skip_validation=skip_validation,
            profile_name=selected_profile or save_profile_name.strip() or None,
            config_profile_name=config_profile_name or None,
            use_batch_api=use_batch_api,
            source_directives=_ui_directives,
        )
        enqueue_job(username, job)
        st.success(f"Queued `{job['memo_name']}`.")

    if "memo_bytes" in st.session_state:
        st.divider()
        unval = st.session_state.get("unvalidated_pages", [])
        if unval:
            st.warning(
                f"Pages {unval} could not be fully validated due to API "
                f"response truncation. Changes on these pages passed without "
                f"QA review. **Manual review is strongly recommended.**"
            )
        st.success("Artifacts are ready for review and download.")
        _manifest_counts = st.session_state.get("manifest", {}).get("counts", {})
        _slides_generated = (
            _manifest_counts.get("slides_inserted", 0)
            + _manifest_counts.get("comp_slides_inserted", 0)
            + _manifest_counts.get("ai_slides_generated", 0)
        )
        metric_cols = st.columns(6)
        metric_cols[0].metric("Applied changes", st.session_state["n_changes"])
        metric_cols[1].metric("Rejected", st.session_state["n_rejected"])
        metric_cols[2].metric("Needs review", st.session_state["n_missed"])
        metric_cols[3].metric("Slides generated", _slides_generated or "—")
        metric_cols[4].metric("Warnings", len(st.session_state.get("warnings", [])))
        _cost_usd = _manifest_counts.get("estimated_cost_microdollars", 0) / 1_000_000
        metric_cols[5].metric("Est. API cost", f"${_cost_usd:.4f}" if _cost_usd else "—")

        # Change type breakdown
        _changes = st.session_state.get("changes", [])
        if _changes:
            _type_counts = {}
            for c in _changes:
                t = c.get("type", "unknown")
                _type_counts[t] = _type_counts.get(t, 0) + 1
            _breakdown = " · ".join(f"{v} {k}" for k, v in sorted(_type_counts.items()))
            st.caption(f"Breakdown: {_breakdown}")

        # Before vs After change report (branded)
        if _changes:
            from app_helpers import build_change_report_html

            with st.expander("Before vs After Report", expanded=True):
                _report_html = build_change_report_html(
                    _changes,
                    manifest=st.session_state.get("manifest"),
                )
                st.markdown(_report_html, unsafe_allow_html=True)

                # Download as standalone HTML
                _full_html = (
                    '<!DOCTYPE html><html><head><meta charset="utf-8">'
                    '<title>Before vs After Report</title>'
                    '<style>body{background:#2b2825;padding:24px;}</style>'
                    '</head><body>' + _report_html + '</body></html>'
                )
                st.download_button(
                    "Download report (HTML)",
                    _full_html.encode("utf-8"),
                    file_name="before_vs_after_report.html",
                    mime="text/html",
                    width="stretch",
                )

        download_cols = st.columns(3)
        download_cols[0].download_button(
            "Download updated memo",
            st.session_state["memo_bytes"],
            file_name=st.session_state["filename"],
            mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            width="stretch",
        )
        download_cols[1].download_button(
            "Download change log",
            st.session_state["log_bytes"],
            file_name="CHANGE_LOG.md",
            mime="text/markdown",
            width="stretch",
        )
        download_cols[2].download_button(
            "Download run manifest",
            st.session_state["manifest_bytes"],
            file_name="run_manifest.json",
            mime="application/json",
            width="stretch",
        )

        warnings = st.session_state.get("warnings", [])
        if warnings:
            with st.expander("Warnings"):
                for warning in warnings:
                    st.warning(f"{warning['stage']}: {warning['message']}")

        with st.expander("Execution log"):
            st.code("\n".join(st.session_state["log_lines"]), language=None)
        st.caption("Run manifest is available as a download artifact.")


def render_history_tab() -> None:
    st.subheader("Run history")
    st.caption("Recent runs, outcomes, approvals, and warning counts for auditing and reruns.")
    try:
        rows = get_recent_runs(None if role == "admin" else username, limit=30)
    except Exception as err:
        st.warning(f"Run history is unavailable: {err}")
        return
    if not rows:
        st.info("No completed or recorded runs yet.")
        return
    st.dataframe(rows, width="stretch", hide_index=True)

    st.divider()
    st.markdown("#### Approval workflow")
    run_options = [row["Run ID"] for row in rows]
    selected_run = st.selectbox("Select run", run_options, index=0)
    details = get_run_details(selected_run)
    if details:
        detail_cols = st.columns(6)
        detail_cols[0].metric("Status", details["status"])
        detail_cols[1].metric("Approval", details["approval_status"])
        detail_cols[2].metric("Changes", details["change_count"])
        detail_cols[3].metric("Warnings", len(details["warnings"]))
        conf = details.get("confidence_score")
        detail_cols[4].metric("Confidence", f"{conf:.0f}/100" if conf is not None else "—")
        detail_cols[5].metric("Slides inserted", details.get("slides_inserted", 0))
        if details["warnings"]:
            with st.expander("Run warnings"):
                for warning in details["warnings"]:
                    st.warning(f"{warning['stage']}: {warning['message']}")
        if conf is not None:
            with st.expander("Accuracy breakdown"):
                acc_cols = st.columns(4)
                acc_cols[0].metric("Confidence", f"{conf:.1f}/100")
                cov = details.get("coverage_pct")
                acc_cols[1].metric("Coverage", f"{cov:.1f}%" if cov is not None else "—")
                corr = details.get("correction_rate_pct")
                acc_cols[2].metric("Correction rate", f"{corr:.1f}%" if corr is not None else "—")
                acc_cols[3].metric("Slides inserted", details.get("slides_inserted", 0))
        with st.form("approval_form"):
            approval_status = st.selectbox(
                "Approval decision",
                ["pending", "approved", "needs_revision", "rejected"],
                index=["pending", "approved", "needs_revision", "rejected"].index(
                    details["approval_status"] if details["approval_status"] in {"pending", "approved", "needs_revision", "rejected"} else "pending"
                ),
            )
            approval_notes = st.text_area("Reviewer notes", value=details["approval_notes"], height=100)
            submitted = st.form_submit_button("Save approval", type="primary")
        if submitted:
            update_run_approval(selected_run, approval_status, username, approval_notes or None)
            st.success(f"Updated approval for `{selected_run}`.")
            st.rerun()
        artifact_paths = get_run_artifact_paths(selected_run)
        if artifact_paths:
            action_cols = st.columns(4)
            memo_path = artifact_paths.get("memo")
            log_path = artifact_paths.get("change_log")
            manifest_path = artifact_paths.get("run_manifest")
            if memo_path and os.path.exists(memo_path):
                action_cols[0].download_button(
                    "Download memo",
                    open(memo_path, "rb").read(),
                    file_name=os.path.basename(memo_path),
                    mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                    width="stretch",
                )
            if log_path and os.path.exists(log_path):
                action_cols[1].download_button(
                    "Download log",
                    open(log_path, "rb").read(),
                    file_name=os.path.basename(log_path),
                    mime="text/markdown",
                    width="stretch",
                )
            if manifest_path and os.path.exists(manifest_path):
                action_cols[2].download_button(
                    "Download manifest",
                    open(manifest_path, "rb").read(),
                    file_name=os.path.basename(manifest_path),
                    mime="application/json",
                    width="stretch",
                )
            if action_cols[3].button("Requeue from history", width="stretch"):
                input_memo_path = artifact_paths.get("input_memo")
                input_proforma_path = artifact_paths.get("input_proforma")
                if input_memo_path and input_proforma_path and os.path.exists(input_memo_path) and os.path.exists(input_proforma_path):
                    payload = {
                        "job_id": uuid.uuid4().hex,
                        "memo_name": details["memo_name"],
                        "memo_bytes": open(input_memo_path, "rb").read(),
                        "proforma_name": details["proforma_name"],
                        "proforma_bytes": open(input_proforma_path, "rb").read(),
                        "schedule_name": os.path.basename(artifact_paths["input_schedule"]) if artifact_paths.get("input_schedule") else None,
                        "schedule_bytes": open(artifact_paths["input_schedule"], "rb").read() if artifact_paths.get("input_schedule") and os.path.exists(artifact_paths["input_schedule"]) else None,
                        "market_data_name": os.path.basename(artifact_paths["input_market_data"]) if artifact_paths.get("input_market_data") else None,
                        "market_data_bytes": open(artifact_paths["input_market_data"], "rb").read() if artifact_paths.get("input_market_data") and os.path.exists(artifact_paths["input_market_data"]) else None,
                        "property_name": details["property_name"] or None,
                        "dry_run": details["dry_run"],
                        "skip_validation": details["skip_validation"],
                        "profile_name": "",
                    }
                    enqueue_job(username, payload)
                    st.success(f"Requeued `{details['memo_name']}`.")
                    st.rerun()
                else:
                    st.error("Stored input artifacts were not found for this run.")


def render_operations_tab() -> None:
    st.subheader("Operations")
    st.caption("Batch execution, platform health, and profile inventory.")
    ops_tabs = st.tabs(["Queue", "Health", "Profiles"])

    with ops_tabs[0]:
        queue = [job for job in get_job_queue(None if role == "admin" else username) if job["status"] in {"queued", "running", "failed"}]
        if queue:
            queue_rows = [
                {
                    "Job ID": item["job_id"],
                    "Status": item["status"],
                    "Memo": item["payload"]["memo_name"],
                    "Proforma": item["payload"]["proforma_name"],
                    "Property": item["payload"].get("property_name") or "",
                    "Preview": "Yes" if item["payload"].get("dry_run") else "No",
                    "Skip QA": "Yes" if item["payload"].get("skip_validation") else "No",
                    "Profile": item["payload"].get("profile_name", ""),
                    "Run ID": item["run_id"],
                    "Error": item["error_message"],
                }
                for item in queue
            ]
            st.dataframe(queue_rows, width="stretch", hide_index=True)
            n_queued = sum(1 for item in queue if item["status"] == "queued")
            n_running = sum(1 for item in queue if item["status"] == "running")
            if n_running:
                st.info(f"Background worker is processing a job. {n_queued} remaining in queue.")
            elif n_queued:
                st.caption(f"{n_queued} queued job(s) will auto-start within ~10 seconds.")
            queue_cols = st.columns(4)
            if queue_cols[0].button("Refresh status", type="primary", width="stretch"):
                st.rerun()
            selected_job_id = queue_cols[1].selectbox(
                "Job",
                [item["job_id"] for item in queue],
                key="ops_job_select",
                label_visibility="collapsed",
            )
            if queue_cols[2].button("Delete selected job", width="stretch"):
                delete_job(selected_job_id)
                st.info(f"Deleted job `{selected_job_id}`.")
                st.rerun()
            if queue_cols[3].button("Retry failed job", width="stretch"):
                failed_job = get_job(selected_job_id)
                if failed_job and failed_job["status"] == "failed":
                    update_job_status(selected_job_id, "queued", error_message=None)
                    st.success(f"Job `{selected_job_id}` moved back to queued.")
                    st.rerun()
                st.warning("Select a failed job to retry.")
        else:
            st.info("No queued jobs yet.")

    with ops_tabs[1]:
        health_rows = get_platform_health()
        st.dataframe(health_rows, width="stretch", hide_index=True)

    with ops_tabs[2]:
        profiles = get_profiles(None if role == "admin" else username)
        if profiles:
            st.dataframe(profiles, width="stretch", hide_index=True)
        else:
            st.info("No saved profiles yet.")


def render_admin_tab() -> None:
    st.subheader("Admin")
    st.caption("Manage users, credits, and review system activity.")
    users = get_users()
    rows = []
    for user_name, user_cfg in users.items():
        credit_limit = int(user_cfg.get("credits_per_week", 5))
        try:
            used_count, remaining_count = get_user_credits(user_name, credit_limit)
        except Exception:
            used_count, remaining_count = 0, 0
        rows.append(
            {
                "User": user_name,
                "Role": user_cfg.get("role", "user"),
                "Used": used_count,
                "Limit": credit_limit,
                "Remaining": remaining_count,
            }
        )
    st.dataframe(rows, width="stretch", hide_index=True)

    admin_tabs = st.tabs(["Invite user", "Add user", "Edit user", "Delete user", "Reset credits", "Invitations", "Recent activity"])

    with admin_tabs[0]:
        st.caption("Generate an invite link to share with a new user.")
        with st.form("invite_user_form"):
            invite_email = st.text_input("Email address")
            invite_role = st.selectbox("Role", ["user", "admin"], index=0)
            invite_credits = st.number_input("Credits per week", min_value=1, value=5)
            invite_submitted = st.form_submit_button("Generate invite link", type="primary")
        if invite_submitted:
            if not invite_email.strip() or "@" not in invite_email:
                st.error("Please enter a valid email address.")
            else:
                token = create_invitation(
                    email=invite_email.strip(),
                    role=invite_role,
                    credits_per_week=int(invite_credits),
                    invited_by=username,
                )
                invite_url = f"https://memochef.streamlit.app?invite={token}"
                st.success(f"Invite created for **{invite_email.strip()}**")
                st.code(invite_url, language=None)
                st.caption("Copy this link and send it to the user.")

    with admin_tabs[1]:
        with st.form("add_user_form"):
            new_username = st.text_input("Username")
            new_password = st.text_input("Password", type="password")
            new_role = st.selectbox("Role", ["user", "admin"], index=0)
            new_credits = st.number_input("Credits per week", min_value=1, value=5)
            submitted = st.form_submit_button("Create user", type="primary")
        if submitted:
            if not new_username.strip():
                st.error("Username is required.")
            elif len(new_password) < 6:
                st.error("Password must be at least 6 characters.")
            else:
                if add_user(new_username.strip(), new_password, new_role, int(new_credits)):
                    st.success(f"User `{new_username}` created.")
                    st.rerun()
                st.error(f"User `{new_username}` already exists.")

    with admin_tabs[2]:
        usernames = [row["User"] for row in rows]
        selected = st.selectbox("User", usernames, index=None, placeholder="Select a user")
        if selected:
            current_cfg = users[selected]
            with st.form("edit_user_form"):
                edit_role = st.selectbox(
                    "Role",
                    ["user", "admin"],
                    index=0 if current_cfg.get("role", "user") == "user" else 1,
                )
                edit_credits = st.number_input(
                    "Credits per week",
                    min_value=1,
                    value=int(current_cfg.get("credits_per_week", 5)),
                )
                edit_password = st.text_input("New password", type="password")
                submitted = st.form_submit_button("Save changes", type="primary")
            if submitted:
                if edit_password and len(edit_password) < 6:
                    st.error("Password must be at least 6 characters.")
                else:
                    update_user(
                        selected,
                        role=edit_role,
                        credits_per_week=int(edit_credits),
                        new_password=edit_password or None,
                    )
                    if selected == username:
                        st.session_state["role"] = edit_role
                        st.session_state["credits_per_week"] = int(edit_credits)
                    st.success(f"Updated `{selected}`.")
                    st.rerun()

    with admin_tabs[3]:
        deletable = [row["User"] for row in rows if row["User"] != username]
        selected = st.selectbox("User to delete", deletable, index=None, placeholder="Select a user")
        if selected and st.button(f"Delete {selected}"):
            delete_user(selected)
            st.success(f"Deleted `{selected}`.")
            st.rerun()

    with admin_tabs[4]:
        selected = st.selectbox(
            "User to reset",
            [row["User"] for row in rows],
            index=None,
            placeholder="Select a user",
        )
        if selected and st.button(f"Reset credits for {selected}"):
            reset_user_credits(selected)
            st.success(f"Credits reset for `{selected}`.")
            st.rerun()

    with admin_tabs[5]:
        invitations = get_invitations()
        if not invitations:
            st.info("No invitations sent yet.")
        else:
            inv_rows = []
            for inv in invitations:
                status = inv["status"]
                if status == "pending" and inv["expires_at"] < datetime.now(inv["expires_at"].tzinfo):
                    status = "expired"
                inv_rows.append({
                    "Email": inv["email"],
                    "Role": inv["role"],
                    "Credits": inv["credits_per_week"],
                    "Status": status,
                    "Invited by": inv["invited_by"],
                    "Sent": inv["created_at"].strftime("%Y-%m-%d %H:%M") if inv["created_at"] else "",
                })
            st.dataframe(inv_rows, width="stretch", hide_index=True)

    with admin_tabs[6]:
        try:
            runs = get_recent_runs(None, limit=30)
        except Exception as err:
            st.warning(f"Recent activity is unavailable: {err}")
        else:
            st.dataframe(runs, width="stretch", hide_index=True)


def render_how_to_tab() -> None:
    st.subheader("How To")
    st.caption("Quick-start guide and tips for your team.")
    with st.expander("Getting Started", expanded=True):
        st.info("Content coming soon")
    with st.expander("Input Requirements"):
        st.info("Content coming soon")
    with st.expander("Understanding Results"):
        st.info("Content coming soon")


with tabs[0]:
    render_new_run_tab()

with tabs[1]:
    render_history_tab()

with tabs[2]:
    render_operations_tab()

with tabs[3]:
    render_how_to_tab()

if role == "admin":
    with tabs[4]:
        render_admin_tab()
