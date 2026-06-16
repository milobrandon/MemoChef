"""
College House SQL — comp & market performance data source.

Pulls monthly property performance (prelease, occupancy, rate per bed,
rate per SF) from Subtext's market research database (Azure SQL,
``StudentResearch``) for use in IC memo comp tables and market
performance slides.

Credentials come from environment variables (populate ``.env`` — see
``.env.example``). The password is REQUIRED; server/database/driver have
defaults matching the standard Subtext research server:

    COLLEGEHOUSE_SQL_SERVER    (default: subtextresearch.database.windows.net)
    COLLEGEHOUSE_SQL_DATABASE  (default: StudentResearch)
    COLLEGEHOUSE_SQL_USERNAME  (required)
    COLLEGEHOUSE_SQL_PASSWORD  (required)
    COLLEGEHOUSE_SQL_DRIVER    (default: ODBC Driver 17 for SQL Server)

The primary table is ``[dbo].[MonthlyPropertyDataByBedAndSF_CH]`` — one row
per property per bedroom count per month, with numerator/denominator pairs
for prelease, occupancy/vacancy, rate, and rate-per-SF.

The server is known to time out routinely, so every public entry point
fails soft: log a warning and return an empty result rather than raising.
A successful pull should always be persisted to an extract workbook so
reruns never depend on the connection.
"""

from __future__ import annotations

import logging
import os
import re
import time

log = logging.getLogger("memo_automator")

DEFAULT_SERVER = "subtextresearch.database.windows.net"
DEFAULT_DATABASE = "StudentResearch"
DEFAULT_DRIVER = "ODBC Driver 17 for SQL Server"
DEFAULT_TIMEOUT_SECONDS = 30
DEFAULT_MONTHS_BACK = 24

MONTHLY_TABLE = "[dbo].[MonthlyPropertyDataByBedAndSF_CH]"
PLANS_TABLE = "[dbo].[Plans]"
PROPERTIES_TABLE = "[dbo].[Properties]"

# Columns pulled for floor-plan detail (current state per plan), SELECT order.
PLAN_COLUMNS = [
    "BuildingName", "PlanName", "Format", "IsStudio",
    "Bedrooms", "Bathrooms", "AreaSF", "Beds",
    "Rate", "RatePerSF", "Occupancy",
]

# Raw columns pulled from the monthly table, in SELECT order.
MONTHLY_COLUMNS = [
    "MonthDate", "BuildingName", "Property_Key", "Bedrooms",
    "InstitutionName", "IPEDS", "TotalBeds",
    "PreleaseNum", "PreleaseDenom",
    "VacNum", "OccNum", "VacDenom",
    "RateNum", "RateDenom", "RateSFNum", "SFDenom",
    "Units", "BedCount",
]

# Derived metrics appended to each row (ratio of the listed num/denom).
DERIVED_METRICS = [
    ("PreleasePct", "PreleaseNum", "PreleaseDenom"),
    ("OccupancyPct", "OccNum", "VacDenom"),
    ("RatePerBed", "RateNum", "RateDenom"),
    ("RatePerSF", "RateSFNum", "SFDenom"),
]


def get_sql_settings() -> dict:
    """Read College House SQL settings from the environment."""
    return {
        "server": os.environ.get("COLLEGEHOUSE_SQL_SERVER", DEFAULT_SERVER),
        "database": os.environ.get("COLLEGEHOUSE_SQL_DATABASE", DEFAULT_DATABASE),
        "username": os.environ.get("COLLEGEHOUSE_SQL_USERNAME", ""),
        "password": os.environ.get("COLLEGEHOUSE_SQL_PASSWORD", ""),
        "driver": os.environ.get("COLLEGEHOUSE_SQL_DRIVER", DEFAULT_DRIVER),
    }


def is_configured() -> bool:
    """True when credentials are present AND pyodbc is importable."""
    settings = get_sql_settings()
    if not (settings["username"] and settings["password"]):
        return False
    try:
        import pyodbc  # noqa: F401
    except ImportError:
        return False
    return True


_DRIVER_FALLBACKS = (
    "ODBC Driver 18 for SQL Server",
    "ODBC Driver 17 for SQL Server",
    "SQL Server",
)


def _resolve_driver(preferred: str, pyodbc_module) -> str:
    """Use the preferred driver if installed, else the best available one."""
    try:
        available = set(pyodbc_module.drivers())
    except Exception:
        return preferred
    if preferred in available:
        return preferred
    for candidate in _DRIVER_FALLBACKS:
        if candidate in available:
            log.info(
                "College House SQL: driver '%s' not installed; using '%s'",
                preferred, candidate,
            )
            return candidate
    return preferred


def _build_connection_string(settings: dict) -> str:
    return (
        f"DRIVER={{{settings['driver']}}};"
        f"SERVER={settings['server']};"
        f"DATABASE={settings['database']};"
        f"UID={settings['username']};"
        f"PWD={settings['password']}"
    )


def _safe_ratio(num, denom):
    try:
        num = float(num)
        denom = float(denom)
    except (TypeError, ValueError):
        return None
    if denom == 0:
        return None
    return num / denom


def _rows_with_derived(cursor_rows, columns: list[str]) -> list[dict]:
    """Convert pyodbc rows to dicts and append derived ratio metrics."""
    out = []
    for raw in cursor_rows:
        row = dict(zip(columns, raw))
        for name, num_col, denom_col in DERIVED_METRICS:
            row[name] = _safe_ratio(row.get(num_col), row.get(denom_col))
        out.append(row)
    return out


def _run_query(
    query: str,
    params: list,
    timeout_seconds: int,
    retries: int,
) -> list | None:
    """Run a query with credential checks, driver resolution, and retries.

    Returns raw cursor rows, or None on any failure (missing config,
    missing pyodbc, connection/login timeouts).
    """
    settings = get_sql_settings()
    if not (settings["username"] and settings["password"]):
        log.warning(
            "College House SQL credentials not configured "
            "(COLLEGEHOUSE_SQL_USERNAME / COLLEGEHOUSE_SQL_PASSWORD). Skipping."
        )
        return None

    try:
        import pyodbc
    except ImportError:
        log.warning("pyodbc not installed; cannot reach College House SQL. Skipping.")
        return None

    settings["driver"] = _resolve_driver(settings["driver"], pyodbc)
    conn_str = _build_connection_string(settings)
    last_error: Exception | None = None
    for attempt in range(retries + 1):
        try:
            log.info(
                "College House SQL: connecting (attempt %d/%d, timeout=%ds)",
                attempt + 1, retries + 1, timeout_seconds,
            )
            conn = pyodbc.connect(conn_str, timeout=timeout_seconds)
            try:
                cursor = conn.cursor()
                cursor.execute(query, params)
                return cursor.fetchall()
            finally:
                conn.close()
        except Exception as e:  # pyodbc errors don't share a clean base class
            last_error = e
            log.warning("College House SQL attempt %d failed: %s", attempt + 1, e)
            if attempt < retries:
                time.sleep(2)

    log.warning(
        "College House SQL unavailable after %d attempt(s): %s. "
        "Continuing without live comp/market data.",
        retries + 1, last_error,
    )
    return None


def fetch_market_performance(
    *,
    institution: str | None = None,
    ipeds: int | None = None,
    property_like: list[str] | None = None,
    months_back: int = DEFAULT_MONTHS_BACK,
    timeout_seconds: int = DEFAULT_TIMEOUT_SECONDS,
    retries: int = 1,
) -> list[dict]:
    """
    Pull monthly comp/market performance rows from College House.

    Filters by institution name (exact match per the DB convention),
    IPEDS id, and/or building-name fragments (OR'd LIKE filters).
    Returns a list of dicts (raw columns + derived PreleasePct,
    OccupancyPct, RatePerBed, RatePerSF), or [] on any failure.
    """
    if not (institution or ipeds or property_like):
        log.warning("College House: no institution/IPEDS/property filter given; skipping pull.")
        return []

    where: list[str] = []
    params: list = []
    if institution:
        where.append("InstitutionName = ?")
        params.append(institution)
    if ipeds:
        where.append("IPEDS = ?")
        params.append(int(ipeds))
    if property_like:
        likes = " OR ".join("BuildingName LIKE ?" for _ in property_like)
        where.append(f"({likes})")
        params.extend(f"%{frag}%" for frag in property_like)
    if months_back and months_back > 0:
        where.append(f"MonthDate >= DATEADD(month, -{int(months_back)}, GETDATE())")

    query = (
        f"SELECT {', '.join(f'[{c}]' for c in MONTHLY_COLUMNS)}\n"
        f"FROM {MONTHLY_TABLE}\n"
        f"WHERE {' AND '.join(where)}\n"
        f"ORDER BY BuildingName, Bedrooms, MonthDate"
    )

    raw = _run_query(query, params, timeout_seconds, retries)
    if raw is None:
        return []
    rows = _rows_with_derived(raw, MONTHLY_COLUMNS)
    log.info("College House SQL: pulled %d monthly rows", len(rows))
    return rows


# Variant designation at the end of a plan name: "4BR/4BA - D1" -> ("D", 1),
# "4X2 Lite B" -> ("B", 0), "Studio - S1" -> ("S", 1).
_VARIANT_RE = re.compile(r"([A-Za-z]{1,3})\s*[-_ ]?\s*(\d{0,3})\s*$")


def _variant_key(plan_name: str) -> tuple[str, int]:
    """Sort key for variant ordering within a bed/bath block — the
    first-named variant (A1 < A2 < B1; A < B) is the base variant."""
    m = _VARIANT_RE.search(str(plan_name or "").strip())
    if not m:
        return ("~", 999)  # unparseable names sort last
    letter = m.group(1).upper()
    number = int(m.group(2)) if m.group(2) else 0
    return (letter, number)


def mark_base_variants(plans: list[dict]) -> list[dict]:
    """Flag the base variant within each (property, bedrooms, bathrooms,
    studio) block: the FIRST-NAMED variant by designation (A1/B1/D1 — letter
    then number). Adds ``IsBaseVariant`` to every plan dict in place."""
    blocks: dict[tuple, list[dict]] = {}
    for plan in plans:
        key = (
            str(plan.get("BuildingName") or ""),
            plan.get("Bedrooms"),
            plan.get("Bathrooms"),
            bool(plan.get("IsStudio")),
        )
        blocks.setdefault(key, []).append(plan)

    for group in blocks.values():
        base = min(
            group,
            key=lambda p: (_variant_key(p.get("PlanName")), str(p.get("PlanName") or "")),
        )
        for plan in group:
            plan["IsBaseVariant"] = plan is base
    return plans


def fetch_floor_plans(
    *,
    institution: str | None = None,
    ipeds: int | None = None,
    property_like: list[str] | None = None,
    timeout_seconds: int = DEFAULT_TIMEOUT_SECONDS,
    retries: int = 1,
) -> list[dict]:
    """
    Pull current floor-plan detail per comp from the ``Plans`` table —
    the only source granular enough for unit-type side-by-side rent rows
    (the monthly table blends all variants of a bedroom count together).

    Institution filtering resolves property keys via the monthly table
    (Plans has no institution column). Returns plan dicts with
    ``IsBaseVariant`` flagged per (property, bed, bath, studio) block,
    or [] on any failure.
    """
    if not (institution or ipeds or property_like):
        return []

    where: list[str] = []
    params: list = []
    if institution or ipeds:
        inst_where, inst_params = [], []
        if institution:
            inst_where.append("InstitutionName = ?")
            inst_params.append(institution)
        if ipeds:
            inst_where.append("IPEDS = ?")
            inst_params.append(int(ipeds))
        where.append(
            f"Pl.propertyKey IN (SELECT DISTINCT Property_Key FROM {MONTHLY_TABLE} "
            f"WHERE {' AND '.join(inst_where)})"
        )
        params.extend(inst_params)
    if property_like:
        likes = " OR ".join("P.[name] LIKE ?" for _ in property_like)
        where.append(f"({likes})")
        params.extend(f"%{frag}%" for frag in property_like)

    query = (
        "SELECT P.[name], Pl.[name], Pl.[format], Pl.[isStudio],\n"
        "       Pl.[bedrooms], Pl.[bathrooms], Pl.[areaSf], Pl.[bedsPurposeBuilt],\n"
        "       Pl.[rate], Pl.[ratePerSf], Pl.[occupancy]\n"
        f"FROM {PLANS_TABLE} Pl\n"
        f"LEFT JOIN {PROPERTIES_TABLE} P ON Pl.propertyKey = P.[key]\n"
        f"WHERE {' AND '.join(where)}\n"
        "ORDER BY P.[name], Pl.[bedrooms], Pl.[bathrooms], Pl.[name]"
    )

    raw = _run_query(query, params, timeout_seconds, retries)
    if raw is None:
        return []
    plans = [dict(zip(PLAN_COLUMNS, r)) for r in raw]
    for plan in plans:
        for col in ("Rate", "RatePerSF", "Occupancy"):
            if plan.get(col) is not None:
                plan[col] = float(plan[col])
    log.info("College House SQL: pulled %d floor plans", len(plans))
    return mark_base_variants(plans)


def _fmt_pct(value) -> str:
    return f"{value:.1%}" if isinstance(value, float) else ""


def _fmt_money(value) -> str:
    return f"${value:,.0f}" if isinstance(value, float) else ""


def _fmt_money2(value) -> str:
    return f"${value:,.2f}" if isinstance(value, float) else ""


def _month_str(value) -> str:
    try:
        return value.strftime("%Y-%m")
    except AttributeError:
        return str(value or "")


def _cycle_start(month) -> tuple[int, int]:
    """(year, month) of the September starting the leasing cycle containing
    ``month`` — student-housing leasing cycles run September → August."""
    return (month.year if month.month >= 9 else month.year - 1, 9)


def _weighted_avg_rate(rows: list[dict]) -> float | None:
    pairs = [
        (r["RatePerBed"], float(r.get("BedCount") or 0))
        for r in rows
        if isinstance(r.get("RatePerBed"), float) and r.get("BedCount")
    ]
    total_w = sum(w for _, w in pairs)
    if not pairs or total_w == 0:
        return None
    return sum(v * w for v, w in pairs) / total_w


def compute_leasing_cycle_rent_growth(rows: list[dict]) -> dict[str, dict]:
    """
    YoY rent growth per property using LEASING-CYCLE AVERAGE rents: the
    bed-weighted average rate per bed over the current cycle (September
    through the property's latest month) vs the same September→month window
    one year prior. Returns {BuildingName: {current_avg_rent,
    prior_avg_rent, rent_growth}} — values are None when either window has
    no rate data.
    """
    by_building: dict[str, list[dict]] = {}
    for row in rows:
        bld = str(row.get("BuildingName") or "")
        if bld and row.get("MonthDate") is not None:
            by_building.setdefault(bld, []).append(row)

    def _in_window(month, start: tuple[int, int], end: tuple[int, int]) -> bool:
        return start <= (month.year, month.month) <= end

    out: dict[str, dict] = {}
    for bld, group in by_building.items():
        latest = max(r["MonthDate"] for r in group)
        start = _cycle_start(latest)
        end = (latest.year, latest.month)
        prior_start = (start[0] - 1, start[1])
        prior_end = (end[0] - 1, end[1])

        cur_avg = _weighted_avg_rate(
            [r for r in group if _in_window(r["MonthDate"], start, end)]
        )
        prior_avg = _weighted_avg_rate(
            [r for r in group if _in_window(r["MonthDate"], prior_start, prior_end)]
        )
        growth = (
            cur_avg / prior_avg - 1
            if isinstance(cur_avg, float) and isinstance(prior_avg, float) and prior_avg != 0
            else None
        )
        out[bld] = {
            "current_avg_rent": cur_avg,
            "prior_avg_rent": prior_avg,
            "rent_growth": growth,
        }
    return out


def summarize_latest_month(rows: list[dict]) -> list[dict]:
    """
    Roll rows up to the latest month per property: total beds, bed-weighted
    prelease/occupancy/rate-per-bed/rate-per-SF across bedroom types, plus
    YoY rent growth computed from leasing-cycle average rents (September →
    latest month vs the same window a year prior).
    """
    latest: dict[str, list[dict]] = {}
    latest_month: dict[str, object] = {}
    for row in rows:
        bld = str(row.get("BuildingName") or "")
        month = row.get("MonthDate")
        if month is None or not bld:
            continue
        if bld not in latest_month or month > latest_month[bld]:
            latest_month[bld] = month
            latest[bld] = [row]
        elif month == latest_month[bld]:
            latest[bld].append(row)

    rent_growth = compute_leasing_cycle_rent_growth(rows)

    summary = []
    for bld in sorted(latest):
        group = latest[bld]

        def _weighted(metric: str, weight_col: str = "BedCount") -> float | None:
            pairs = [
                (r[metric], float(r.get(weight_col) or 0))
                for r in group
                if isinstance(r.get(metric), float) and r.get(weight_col)
            ]
            total_w = sum(w for _, w in pairs)
            if not pairs or total_w == 0:
                return None
            return sum(v * w for v, w in pairs) / total_w

        summary.append({
            "BuildingName": bld,
            "InstitutionName": group[0].get("InstitutionName"),
            "MonthDate": latest_month[bld],
            "TotalBeds": sum(int(r.get("BedCount") or 0) for r in group),
            "PreleasePct": _weighted("PreleasePct"),
            "OccupancyPct": _weighted("OccupancyPct"),
            "RatePerBed": _weighted("RatePerBed"),
            "RatePerSF": _weighted("RatePerSF"),
            "RentGrowthYoY": rent_growth.get(bld, {}).get("rent_growth"),
        })
    return summary


def format_market_performance_text(rows: list[dict], plans: list[dict] | None = None) -> str:
    """
    Render pulled rows (and optional floor-plan detail) as compact text in
    the same tab-delimited shape that ``extract_market_data`` emits, so the
    result can be appended to (or used instead of) market-workbook text and
    flow through the existing market mapping prompts unchanged.
    """
    if not rows and not plans:
        return ""

    sections = []
    rows = rows or []
    plans = plans or []

    # Section 1: latest-month comp summary (the comp-table refresh view).
    summary = summarize_latest_month(rows) if rows else []
    lines = [
        f"\n{'=' * 70}",
        "TAB: Comp Performance Summary (latest month per property)",
        f"{'=' * 70}",
        "Note: YoY Rent Growth uses LEASING-CYCLE AVERAGE rents — the "
        "bed-weighted average rate per bed from September through the latest "
        "month, vs the same September-to-month window one year prior.",
        "Row 1:\tProperty\tInstitution\tMonth\tTotal Beds\tPrelease %\tOccupancy %\tRate/Bed\tRate/SF\tYoY Rent Growth",
    ]
    for i, s in enumerate(summary, start=2):
        lines.append(
            f"Row {i}:\t{s['BuildingName']}\t{s['InstitutionName'] or ''}\t"
            f"{_month_str(s['MonthDate'])}\t{s['TotalBeds']}\t"
            f"{_fmt_pct(s['PreleasePct'])}\t{_fmt_pct(s['OccupancyPct'])}\t"
            f"{_fmt_money(s['RatePerBed'])}\t{_fmt_money2(s['RatePerSF'])}\t"
            f"{_fmt_pct(s['RentGrowthYoY'])}"
        )
    if summary:
        sections.append("\n".join(lines))

    # Section 2: monthly time series per property/bedroom (trend view).
    lines = [
        f"\n{'=' * 70}",
        "TAB: Monthly Performance By Property And Bedroom",
        f"{'=' * 70}",
        "Row 1:\tProperty\tBedrooms\tMonth\tBeds\tPrelease %\tOccupancy %\tRate/Bed\tRate/SF",
    ]
    for i, r in enumerate(rows, start=2):
        lines.append(
            f"Row {i}:\t{r.get('BuildingName') or ''}\t{r.get('Bedrooms') if r.get('Bedrooms') is not None else ''}\t"
            f"{_month_str(r.get('MonthDate'))}\t{r.get('BedCount') if r.get('BedCount') is not None else ''}\t"
            f"{_fmt_pct(r.get('PreleasePct'))}\t{_fmt_pct(r.get('OccupancyPct'))}\t"
            f"{_fmt_money(r.get('RatePerBed'))}\t{_fmt_money2(r.get('RatePerSF'))}"
        )
    if rows:
        sections.append("\n".join(lines))

    # Section 3: floor-plan detail (the ONLY valid source for unit-type
    # side-by-side rent rows — monthly data blends variants together).
    if plans:
        lines = [
            f"\n{'=' * 70}",
            "TAB: Floor Plan Detail (current, per plan)",
            f"{'=' * 70}",
            "Note: Use THIS tab for unit-type side-by-side comp rents. "
            "'Base Variant' = the first-named variant (A1/B1/D1) within the "
            "property's bed/bath block; when the run brief says base-variant "
            "rents only, use that row's rent (never a range or average).",
            "Row 1:\tProperty\tPlan\tStudio\tBed\tBath\tSF\tBeds\tRent/Bed\tRent/SF\tOccupancy\tBase Variant",
        ]
        for i, p in enumerate(plans, start=2):
            lines.append(
                f"Row {i}:\t{p.get('BuildingName') or ''}\t{p.get('PlanName') or ''}\t"
                f"{'Yes' if p.get('IsStudio') else 'No'}\t"
                f"{p.get('Bedrooms') if p.get('Bedrooms') is not None else ''}\t"
                f"{p.get('Bathrooms') if p.get('Bathrooms') is not None else ''}\t"
                f"{p.get('AreaSF') if p.get('AreaSF') is not None else ''}\t"
                f"{p.get('Beds') if p.get('Beds') is not None else ''}\t"
                f"{_fmt_money(p.get('Rate'))}\t{_fmt_money2(p.get('RatePerSF'))}\t"
                f"{_fmt_pct(p.get('Occupancy'))}\t"
                f"{'BASE' if p.get('IsBaseVariant') else ''}"
            )
        sections.append("\n".join(lines))

    header = [
        f"\n{'=' * 70}",
        "MARKET DATA (from College House SQL — StudentResearch)",
        f"{'=' * 70}",
    ]
    return "\n".join(header) + "\n" + "\n\n".join(sections)


def write_extract_workbook(
    rows: list[dict], path: str, plans: list[dict] | None = None
) -> str | None:
    """
    Persist a successful pull to an .xlsx extract (latest-month summary,
    raw monthly rows, and floor-plan detail when supplied) so reruns and
    the managed agent never depend on the connection.
    Returns the path written, or None when there is nothing to write.
    """
    if not rows and not plans:
        return None
    rows = rows or []

    import openpyxl

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Comp Performance Summary"
    ws.append([
        "Property", "Institution", "Month", "Total Beds",
        "Prelease %", "Occupancy %", "Rate Per Bed", "Rate Per SF",
        "YoY Rent Growth (leasing-cycle avg, Sep-to-date)",
    ])
    for s in summarize_latest_month(rows):
        ws.append([
            s["BuildingName"], s["InstitutionName"], _month_str(s["MonthDate"]),
            s["TotalBeds"], s["PreleasePct"], s["OccupancyPct"],
            s["RatePerBed"], s["RatePerSF"], s["RentGrowthYoY"],
        ])

    raw = wb.create_sheet("Monthly Raw Data")
    derived_names = [name for name, _, _ in DERIVED_METRICS]
    raw.append(MONTHLY_COLUMNS + derived_names)
    for r in rows:
        raw.append(
            [_month_str(r.get("MonthDate"))]
            + [r.get(c) for c in MONTHLY_COLUMNS[1:]]
            + [r.get(n) for n in derived_names]
        )

    if plans:
        fp = wb.create_sheet("Floor Plan Detail")
        fp.append(PLAN_COLUMNS + ["IsBaseVariant"])
        for p in plans:
            fp.append([p.get(c) for c in PLAN_COLUMNS] + [bool(p.get("IsBaseVariant"))])

    wb.save(path)
    log.info("College House SQL: extract written to %s", path)
    return path


def extract_college_house_market_data(cfg: dict, output_dir: str | None = None) -> str:
    """
    Top-level pipeline hook: pull comp/market performance per the
    ``market_data.college_house`` config section and return compact text
    (same contract as ``extract_market_data`` — empty string on any failure).

    Config shape (config.yaml)::

        market_data:
          college_house:
            institution: "University of Central Florida"  # or ipeds / properties
            ipeds: 132903
            properties: ["Hub Orlando", "Verve Orlando"]
            months_back: 24
            timeout_seconds: 30

    When ``output_dir`` is given, a successful pull is also persisted to
    ``college_house_extract.xlsx`` in that directory.
    """
    ch_cfg = (cfg.get("market_data", {}) or {}).get("college_house", {}) or {}
    institution = ch_cfg.get("institution")
    ipeds = ch_cfg.get("ipeds")
    properties = ch_cfg.get("properties") or None
    if not (institution or ipeds or properties):
        return ""

    timeout = ch_cfg.get("timeout_seconds", DEFAULT_TIMEOUT_SECONDS)
    rows = fetch_market_performance(
        institution=institution,
        ipeds=ipeds,
        property_like=properties,
        months_back=ch_cfg.get("months_back", DEFAULT_MONTHS_BACK),
        timeout_seconds=timeout,
    )
    if not rows:
        return ""

    plans = fetch_floor_plans(
        institution=institution,
        ipeds=ipeds,
        property_like=properties,
        timeout_seconds=timeout,
    )

    if output_dir:
        try:
            write_extract_workbook(
                rows, os.path.join(output_dir, "college_house_extract.xlsx"),
                plans=plans,
            )
        except Exception as e:
            log.warning("College House SQL: failed to write extract workbook: %s", e)

    return format_market_performance_text(rows, plans=plans)
