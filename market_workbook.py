"""Market Data Workbook generator for Memo Chef.

Produces a multi-tab Excel (.xlsx) workbook with formatted tables and charts
for rent comps, sale comps, submarket stats, macro indicators, and unit mix.

Usage::

    from market_workbook import generate_workbook, load_sample_data

    data = load_sample_data("sample_market_data.json")
    path = generate_workbook(data, "output/Market_Data.xlsx")
"""

from __future__ import annotations

import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart, Series
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Brand colours (Subtext theme)
# ---------------------------------------------------------------------------
_DARK_GREEN = "16352E"
_LIME = "C1D100"
_CREAM = "F7F1E3"
_RUST = "A95818"
_MAHOGANY = "512213"
_BROWN = "2B2825"
_WHITE = "FFFFFF"

_CHART_COLORS = [_DARK_GREEN, _LIME, _RUST, _MAHOGANY, _BROWN, _CREAM]

# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclass
class PropertyInfo:
    name: str
    address: str
    city: str
    state: str
    submarket: str
    msa: str
    property_type: str  # multifamily, mixed-use, etc.


@dataclass
class RentComp:
    name: str
    address: str
    distance_mi: float
    units: int
    year_built: int
    occupancy_pct: float
    unit_rents: dict[str, float]  # {"Studio": 1450, "1BR": 1800, ...}
    concessions: str
    as_of_date: str
    source: str


@dataclass
class SaleComp:
    name: str
    address: str
    sale_date: str
    sale_price: float
    units: int
    price_per_unit: float
    cap_rate: float
    source: str


@dataclass
class SubmarketStats:
    name: str
    avg_rent: float
    vacancy_pct: float
    absorption_units: int
    pipeline_units: int
    rent_growth_yoy_pct: float
    median_hh_income: float
    population: int
    as_of_date: str
    source: str


@dataclass
class MacroEntry:
    date: str
    sofr_rate: float
    treasury_10yr: float
    cpi_yoy: float
    unemployment: float
    source: str


@dataclass
class UnitMixEntry:
    unit_type: str
    unit_count: int
    avg_sf: float
    avg_rent: float
    rent_per_sf: float


@dataclass
class WorkbookMetadata:
    prepared_by: str
    prepared_date: str
    property_name: str
    notes: str = ""


@dataclass
class MarketDataWorkbook:
    property: PropertyInfo
    rent_comps: list[RentComp] = field(default_factory=list)
    sale_comps: list[SaleComp] = field(default_factory=list)
    submarket: SubmarketStats | None = None
    macro: list[MacroEntry] = field(default_factory=list)
    unit_mix: list[UnitMixEntry] = field(default_factory=list)
    metadata: WorkbookMetadata | None = None


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------
_HEADER_FONT = Font(name="Pragmatica Bold", bold=True, color=_WHITE, size=11)
_HEADER_FILL = PatternFill(start_color=_DARK_GREEN, end_color=_DARK_GREEN, fill_type="solid")
_BODY_FONT = Font(name="Pragmatica Book", size=10)
_TITLE_FONT = Font(name="Pragmatica Bold", bold=True, size=16, color=_DARK_GREEN)
_SUBTITLE_FONT = Font(name="Pragmatica Book", size=12, color=_BROWN)
_THIN_BORDER = Border(
    left=Side(style="thin", color=_CREAM),
    right=Side(style="thin", color=_CREAM),
    top=Side(style="thin", color=_CREAM),
    bottom=Side(style="thin", color=_CREAM),
)


def _apply_header_row(ws, row: int, col_count: int) -> None:
    """Format a header row with brand styling."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _apply_body_formatting(ws, start_row: int, end_row: int, col_count: int) -> None:
    """Apply body font and alternating row shading."""
    cream_fill = PatternFill(start_color=_CREAM, end_color=_CREAM, fill_type="solid")
    for row in range(start_row, end_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if (row - start_row) % 2 == 1:
                cell.fill = cream_fill


def _auto_column_widths(ws, min_width: int = 14) -> None:
    """Set column widths based on content, with a minimum."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 4, 30))


def _set_chart_style(chart, title: str) -> None:
    """Apply consistent chart styling."""
    chart.title = title
    chart.style = 10
    chart.width = 18
    chart.height = 12


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_cover_sheet(wb: Workbook, data: MarketDataWorkbook) -> None:
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_properties.tabColor = _DARK_GREEN

    ws.merge_cells("B3:F3")
    ws["B3"] = "Market Data Workbook"
    ws["B3"].font = Font(name="Pragmatica Bold", bold=True, size=24, color=_DARK_GREEN)

    rows = [
        ("Property:", data.property.name),
        ("Address:", f"{data.property.address}, {data.property.city}, {data.property.state}"),
        ("Submarket:", data.property.submarket),
        ("MSA:", data.property.msa),
        ("Type:", data.property.property_type),
    ]
    if data.metadata:
        rows += [
            ("Prepared by:", data.metadata.prepared_by),
            ("Date:", data.metadata.prepared_date),
        ]

    for i, (label, value) in enumerate(rows, start=5):
        ws.cell(row=i, column=2, value=label).font = Font(
            name="Pragmatica Bold", bold=True, size=12, color=_BROWN,
        )
        ws.cell(row=i, column=3, value=value).font = _SUBTITLE_FONT

    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 50


def _write_rent_comps(wb: Workbook, comps: list[RentComp]) -> None:
    if not comps:
        return
    ws = wb.create_sheet("Rent Comps")
    ws.sheet_properties.tabColor = _LIME

    headers = ["Property", "Address", "Distance (mi)", "Units", "Year Built",
               "Occupancy", "Avg Rent", "Concessions", "As-Of", "Source"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_row(ws, 1, len(headers))

    for i, c in enumerate(comps, start=2):
        avg_rent = sum(c.unit_rents.values()) / len(c.unit_rents) if c.unit_rents else 0
        row_data = [c.name, c.address, c.distance_mi, c.units, c.year_built,
                    c.occupancy_pct, avg_rent, c.concessions, c.as_of_date, c.source]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val)

    end_row = len(comps) + 1
    _apply_body_formatting(ws, 2, end_row, len(headers))

    # Format occupancy as percentage
    for row in range(2, end_row + 1):
        ws.cell(row=row, column=6).number_format = '0.0%'
    # Format rent as currency
    for row in range(2, end_row + 1):
        ws.cell(row=row, column=7).number_format = '$#,##0'

    _auto_column_widths(ws)
    ws.freeze_panes = "A2"

    # Bar chart: Avg Rent by Comp
    if len(comps) >= 2:
        chart = BarChart()
        _set_chart_style(chart, "Asking Rent by Comp")
        chart.y_axis.title = "Avg Rent ($)"
        cats = Reference(ws, min_col=1, min_row=2, max_row=end_row)
        vals = Reference(ws, min_col=7, min_row=1, max_row=end_row)
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws.add_chart(chart, "L2")

    # Scatter chart: Rent vs Distance
    if len(comps) >= 2:
        scatter = ScatterChart()
        _set_chart_style(scatter, "Rent vs Distance")
        scatter.x_axis.title = "Distance (mi)"
        scatter.y_axis.title = "Avg Rent ($)"
        x_vals = Reference(ws, min_col=3, min_row=2, max_row=end_row)
        y_vals = Reference(ws, min_col=7, min_row=2, max_row=end_row)
        series = Series(y_vals, x_vals, title="Rent vs Distance")
        scatter.series.append(series)
        ws.add_chart(scatter, "L18")


def _write_sale_comps(wb: Workbook, comps: list[SaleComp]) -> None:
    if not comps:
        return
    ws = wb.create_sheet("Sale Comps")
    ws.sheet_properties.tabColor = _RUST

    headers = ["Property", "Address", "Sale Date", "Sale Price", "Units",
               "$/Unit", "Cap Rate", "Source"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_row(ws, 1, len(headers))

    for i, c in enumerate(comps, start=2):
        row_data = [c.name, c.address, c.sale_date, c.sale_price, c.units,
                    c.price_per_unit, c.cap_rate, c.source]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val)

    end_row = len(comps) + 1
    _apply_body_formatting(ws, 2, end_row, len(headers))

    for row in range(2, end_row + 1):
        ws.cell(row=row, column=4).number_format = '$#,##0'
        ws.cell(row=row, column=6).number_format = '$#,##0'
        ws.cell(row=row, column=7).number_format = '0.00%'

    _auto_column_widths(ws)
    ws.freeze_panes = "A2"

    # Bar chart: $/Unit by Comp
    if len(comps) >= 2:
        chart = BarChart()
        _set_chart_style(chart, "Price per Unit by Comp")
        chart.y_axis.title = "$/Unit"
        cats = Reference(ws, min_col=1, min_row=2, max_row=end_row)
        vals = Reference(ws, min_col=6, min_row=1, max_row=end_row)
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "J2")

    # Line chart: Cap Rate over time
    if len(comps) >= 2:
        line = LineChart()
        _set_chart_style(line, "Cap Rate Trend")
        line.y_axis.title = "Cap Rate"
        line.y_axis.numFmt = '0.00%'
        cats = Reference(ws, min_col=3, min_row=2, max_row=end_row)
        vals = Reference(ws, min_col=7, min_row=1, max_row=end_row)
        line.add_data(vals, titles_from_data=True)
        line.set_categories(cats)
        ws.add_chart(line, "J18")


def _write_submarket(wb: Workbook, stats: SubmarketStats) -> None:
    ws = wb.create_sheet("Submarket Overview")
    ws.sheet_properties.tabColor = _DARK_GREEN

    # Key stats table
    metrics = [
        ("Submarket", stats.name),
        ("Avg Rent", stats.avg_rent),
        ("Vacancy Rate", stats.vacancy_pct),
        ("Net Absorption (units)", stats.absorption_units),
        ("Pipeline (units)", stats.pipeline_units),
        ("Rent Growth YoY", stats.rent_growth_yoy_pct),
        ("Median HH Income", stats.median_hh_income),
        ("Population", stats.population),
        ("As-of Date", stats.as_of_date),
        ("Source", stats.source),
    ]

    headers = ["Metric", "Value"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_row(ws, 1, 2)

    for i, (metric, value) in enumerate(metrics, start=2):
        ws.cell(row=i, column=1, value=metric)
        ws.cell(row=i, column=2, value=value)

    end_row = len(metrics) + 1
    _apply_body_formatting(ws, 2, end_row, 2)

    # Format specific rows
    ws.cell(row=3, column=2).number_format = '$#,##0'
    ws.cell(row=4, column=2).number_format = '0.0%'
    ws.cell(row=7, column=2).number_format = '0.0%'
    ws.cell(row=8, column=2).number_format = '$#,##0'
    ws.cell(row=9, column=2).number_format = '#,##0'

    _auto_column_widths(ws)

    # Pie chart: Occupied vs Vacant
    ws.cell(row=1, column=5, value="Segment")
    ws.cell(row=1, column=6, value="Pct")
    ws.cell(row=2, column=5, value="Occupied")
    ws.cell(row=2, column=6, value=1.0 - stats.vacancy_pct)
    ws.cell(row=3, column=5, value="Vacant")
    ws.cell(row=3, column=6, value=stats.vacancy_pct)
    _apply_header_row(ws, 1, 6)

    pie = PieChart()
    _set_chart_style(pie, "Occupancy Split")
    cats = Reference(ws, min_col=5, min_row=2, max_row=3)
    vals = Reference(ws, min_col=6, min_row=1, max_row=3)
    pie.add_data(vals, titles_from_data=True)
    pie.set_categories(cats)
    ws.add_chart(pie, "D14")


def _write_macro(wb: Workbook, entries: list[MacroEntry]) -> None:
    if not entries:
        return
    ws = wb.create_sheet("Macro Indicators")
    ws.sheet_properties.tabColor = _MAHOGANY

    headers = ["Date", "SOFR", "10yr Treasury", "CPI YoY", "Unemployment", "Source"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_row(ws, 1, len(headers))

    for i, e in enumerate(entries, start=2):
        row_data = [e.date, e.sofr_rate, e.treasury_10yr, e.cpi_yoy, e.unemployment, e.source]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val)

    end_row = len(entries) + 1
    _apply_body_formatting(ws, 2, end_row, len(headers))

    for row in range(2, end_row + 1):
        for col in [2, 3, 4, 5]:
            ws.cell(row=row, column=col).number_format = '0.00%'

    _auto_column_widths(ws)
    ws.freeze_panes = "A2"

    # Line chart: Rate Trends (SOFR + 10yr)
    if len(entries) >= 2:
        line = LineChart()
        _set_chart_style(line, "Interest Rate Trends")
        line.y_axis.title = "Rate"
        line.y_axis.numFmt = '0.00%'
        cats = Reference(ws, min_col=1, min_row=2, max_row=end_row)
        for col_idx, title in [(2, "SOFR"), (3, "10yr Treasury")]:
            vals = Reference(ws, min_col=col_idx, min_row=1, max_row=end_row)
            line.add_data(vals, titles_from_data=True)
        line.set_categories(cats)
        ws.add_chart(line, "H2")

    # Line chart: CPI Trend
    if len(entries) >= 2:
        cpi_line = LineChart()
        _set_chart_style(cpi_line, "CPI YoY Trend")
        cpi_line.y_axis.title = "CPI YoY"
        cpi_line.y_axis.numFmt = '0.00%'
        cats = Reference(ws, min_col=1, min_row=2, max_row=end_row)
        vals = Reference(ws, min_col=4, min_row=1, max_row=end_row)
        cpi_line.add_data(vals, titles_from_data=True)
        cpi_line.set_categories(cats)
        ws.add_chart(cpi_line, "H18")


def _write_unit_mix(wb: Workbook, units: list[UnitMixEntry]) -> None:
    if not units:
        return
    ws = wb.create_sheet("Unit Mix Summary")
    ws.sheet_properties.tabColor = _LIME

    headers = ["Unit Type", "Count", "Avg SF", "Avg Rent", "Rent/SF"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_row(ws, 1, len(headers))

    for i, u in enumerate(units, start=2):
        row_data = [u.unit_type, u.unit_count, u.avg_sf, u.avg_rent, u.rent_per_sf]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val)

    # Totals row
    total_row = len(units) + 2
    ws.cell(row=total_row, column=1, value="Total")
    ws.cell(row=total_row, column=1).font = Font(name="Pragmatica Bold", bold=True, size=10)
    ws.cell(row=total_row, column=2, value=sum(u.unit_count for u in units))
    total_count = sum(u.unit_count for u in units)
    if total_count > 0:
        weighted_sf = sum(u.avg_sf * u.unit_count for u in units) / total_count
        weighted_rent = sum(u.avg_rent * u.unit_count for u in units) / total_count
        ws.cell(row=total_row, column=3, value=round(weighted_sf, 0))
        ws.cell(row=total_row, column=4, value=round(weighted_rent, 0))
        ws.cell(row=total_row, column=5, value=round(weighted_rent / weighted_sf, 2) if weighted_sf else 0)

    end_row = len(units) + 1
    _apply_body_formatting(ws, 2, end_row, len(headers))

    for row in range(2, total_row + 1):
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=4).number_format = '$#,##0'
        ws.cell(row=row, column=5).number_format = '$#0.00'

    _auto_column_widths(ws)
    ws.freeze_panes = "A2"

    # Bar chart: Unit Count by Type
    if len(units) >= 2:
        chart = BarChart()
        _set_chart_style(chart, "Unit Count by Type")
        chart.y_axis.title = "Units"
        cats = Reference(ws, min_col=1, min_row=2, max_row=end_row)
        vals = Reference(ws, min_col=2, min_row=1, max_row=end_row)
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "G2")

    # Bar chart: Avg Rent by Type
    if len(units) >= 2:
        chart2 = BarChart()
        _set_chart_style(chart2, "Avg Rent by Unit Type")
        chart2.y_axis.title = "Rent ($)"
        cats = Reference(ws, min_col=1, min_row=2, max_row=end_row)
        vals = Reference(ws, min_col=4, min_row=1, max_row=end_row)
        chart2.add_data(vals, titles_from_data=True)
        chart2.set_categories(cats)
        ws.add_chart(chart2, "G18")


def _write_sources(wb: Workbook, data: MarketDataWorkbook) -> None:
    ws = wb.create_sheet("Sources & Notes")
    ws.sheet_properties.tabColor = _BROWN

    ws.cell(row=1, column=1, value="Sources & Methodology")
    ws.cell(row=1, column=1).font = _TITLE_FONT

    notes = [
        f"Property: {data.property.name}",
        f"Prepared: {data.metadata.prepared_date if data.metadata else 'N/A'}",
        f"Prepared by: {data.metadata.prepared_by if data.metadata else 'N/A'}",
        "",
        "Data Sources:",
    ]

    sources_seen = set()
    for comp in data.rent_comps:
        sources_seen.add(comp.source)
    for comp in data.sale_comps:
        sources_seen.add(comp.source)
    if data.submarket:
        sources_seen.add(data.submarket.source)
    for entry in data.macro:
        sources_seen.add(entry.source)

    for s in sorted(sources_seen):
        notes.append(f"  - {s}")

    notes += [
        "",
        "Disclaimers:",
        "  - Market data is provided for informational purposes only.",
        "  - Verify all data points against primary sources before use in IC materials.",
        "  - Rent and sale comp data may be subject to licensing restrictions.",
    ]

    if data.metadata and data.metadata.notes:
        notes += ["", "Notes:", data.metadata.notes]

    for i, line in enumerate(notes, start=3):
        ws.cell(row=i, column=1, value=line).font = _BODY_FONT

    ws.column_dimensions["A"].width = 80


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_workbook(data: MarketDataWorkbook, output_path: str) -> str:
    """Generate the full market data Excel workbook.

    Returns the output file path.
    """
    wb = Workbook()

    _write_cover_sheet(wb, data)
    _write_rent_comps(wb, data.rent_comps)
    _write_sale_comps(wb, data.sale_comps)
    if data.submarket:
        _write_submarket(wb, data.submarket)
    _write_macro(wb, data.macro)
    _write_unit_mix(wb, data.unit_mix)
    _write_sources(wb, data)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Market data workbook saved to %s", output_path)
    return output_path


def load_sample_data(json_path: str) -> MarketDataWorkbook:
    """Load market data from a JSON file into dataclass structures."""
    with open(json_path) as f:
        raw = json.load(f)

    prop = PropertyInfo(**raw["property"])

    rent_comps = [RentComp(**c) for c in raw.get("rent_comps", [])]
    sale_comps = [SaleComp(**c) for c in raw.get("sale_comps", [])]

    submarket = None
    if "submarket" in raw:
        submarket = SubmarketStats(**raw["submarket"])

    macro = [MacroEntry(**e) for e in raw.get("macro", [])]
    unit_mix = [UnitMixEntry(**u) for u in raw.get("unit_mix", [])]

    metadata = None
    if "metadata" in raw:
        metadata = WorkbookMetadata(**raw["metadata"])

    return MarketDataWorkbook(
        property=prop,
        rent_comps=rent_comps,
        sale_comps=sale_comps,
        submarket=submarket,
        macro=macro,
        unit_mix=unit_mix,
        metadata=metadata,
    )
