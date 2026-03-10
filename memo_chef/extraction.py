"""Supplemental data extraction for PDF, URL, Excel, and CSV sources."""
from __future__ import annotations

import csv
import logging

log = logging.getLogger(__name__)


def extract_supplemental(source: str, source_type: str) -> str:
    """Extract text from a supplemental data source.

    Args:
        source: File path or URL.
        source_type: One of "pdf", "url", "excel", "csv".

    Returns:
        Plain text representation of the source data.
    """
    extractors = {
        "pdf": _extract_pdf,
        "url": _extract_url,
        "excel": _extract_excel,
        "csv": _extract_csv,
    }
    extractor = extractors.get(source_type)
    if extractor is None:
        raise ValueError(f"Unsupported source_type: {source_type!r}")
    return extractor(source)


def _extract_pdf(path: str) -> str:
    """Extract text and tables from a PDF using pdfplumber."""
    import pdfplumber

    parts: list[str] = []
    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            if text.strip():
                parts.append(f"--- Page {i + 1} ---")
                parts.append(text.strip())

            tables = page.extract_tables() or []
            for t_idx, table in enumerate(tables):
                parts.append(f"Table {t_idx + 1}:")
                for row in table:
                    cells = [str(c) if c is not None else "" for c in row]
                    parts.append("\t".join(cells))
    return "\n".join(parts)


def _extract_url(url: str) -> str:
    """Extract visible text from a URL using requests + BeautifulSoup."""
    import requests
    from bs4 import BeautifulSoup

    resp = requests.get(url, timeout=30, headers={"User-Agent": "MemoChef/1.0"})
    resp.raise_for_status()

    soup = BeautifulSoup(resp.text, "html.parser")

    for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
        tag.decompose()

    text = soup.get_text(separator="\n", strip=True)
    lines = [line for line in text.splitlines() if line.strip()]
    return "\n".join(lines)


def _extract_excel(path: str) -> str:
    """Extract all sheets from an Excel file as tab-delimited text."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text: list[str] = []
        for row in ws.iter_rows(max_row=250, max_col=20, values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c for c in cells):
                rows_text.append("\t".join(cells))
        if rows_text:
            parts.append(f"TAB: {sheet_name}")
            parts.extend(rows_text)
            parts.append("")
    wb.close()
    return "\n".join(parts)


def _extract_csv(path: str) -> str:
    """Extract CSV file as tab-delimited text."""
    parts: list[str] = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row in reader:
            parts.append("\t".join(row))
    return "\n".join(parts)


def extract_comp_urls(comp_urls: list[dict]) -> str:
    """Scrape competitive property websites and return formatted text.

    Args:
        comp_urls: List of dicts with keys ``url``, ``label``, ``guidance``.

    Returns:
        Combined scraped text with headers per property.
    """
    sections: list[str] = []
    for entry in comp_urls:
        url = entry.get("url", "").strip()
        label = entry.get("label", "").strip() or url
        guidance = entry.get("guidance", "").strip()
        if not url:
            continue
        header = f"=== COMP PROPERTY: {label} ==="
        parts = [header, f"Source URL: {url}"]
        if guidance:
            parts.append(f"User guidance: {guidance}")
        try:
            text = _extract_url(url)
            parts.append(text)
        except Exception as exc:
            log.warning("Failed to scrape comp URL %s: %s", url, exc)
            parts.append(f"[Scraping failed: {exc}]")
        sections.append("\n".join(parts))
    return "\n\n".join(sections)
