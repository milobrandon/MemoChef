"""Tests for college_house.py — College House SQL comp/market performance.

All tests run without a real database: pyodbc is faked via sys.modules.
"""

from __future__ import annotations

import datetime
import sys
import types

import openpyxl
import pytest

import college_house
from college_house import (
    MONTHLY_COLUMNS,
    _safe_ratio,
    compute_leasing_cycle_rent_growth,
    extract_college_house_market_data,
    fetch_market_performance,
    format_market_performance_text,
    is_configured,
    summarize_latest_month,
    write_extract_workbook,
)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _raw_row(
    month, building="Hub Orlando", bedrooms=4, beds=100,
    prelease=(90, 100), occ=(95, 100), rate=(900, 1), rate_sf=(3, 1),
):
    """Build a raw tuple in MONTHLY_COLUMNS order."""
    values = {
        "MonthDate": month,
        "BuildingName": building,
        "Property_Key": 1,
        "Bedrooms": bedrooms,
        "InstitutionName": "University of Central Florida",
        "IPEDS": 132903,
        "TotalBeds": beds,
        "PreleaseNum": prelease[0], "PreleaseDenom": prelease[1],
        "VacNum": occ[1] - occ[0], "OccNum": occ[0], "VacDenom": occ[1],
        "RateNum": rate[0], "RateDenom": rate[1],
        "RateSFNum": rate_sf[0], "SFDenom": rate_sf[1],
        "Units": beds // bedrooms, "BedCount": beds,
    }
    return tuple(values[c] for c in MONTHLY_COLUMNS)


class _FakeCursor:
    def __init__(self, rows):
        self._rows = rows
        self.executed = None
        self.params = None

    def execute(self, query, params):
        self.executed = query
        self.params = params

    def fetchall(self):
        return self._rows


class _FakeConn:
    def __init__(self, rows):
        self._cursor = _FakeCursor(rows)
        self.closed = False

    def cursor(self):
        return self._cursor

    def close(self):
        self.closed = True


@pytest.fixture
def fake_pyodbc(monkeypatch):
    """Install a fake pyodbc into sys.modules; yields the module for tweaks."""
    mod = types.ModuleType("pyodbc")
    mod._rows = []
    mod._fail_times = 0
    mod._connections = []

    def connect(conn_str, timeout=30):
        if mod._fail_times > 0:
            mod._fail_times -= 1
            raise OSError("login timeout expired")
        conn = _FakeConn(mod._rows)
        mod._connections.append(conn)
        return conn

    mod.connect = connect
    monkeypatch.setitem(sys.modules, "pyodbc", mod)
    return mod


@pytest.fixture
def creds(monkeypatch):
    monkeypatch.setenv("COLLEGEHOUSE_SQL_USERNAME", "testuser")
    monkeypatch.setenv("COLLEGEHOUSE_SQL_PASSWORD", "testpass")


@pytest.fixture
def no_creds(monkeypatch):
    monkeypatch.delenv("COLLEGEHOUSE_SQL_USERNAME", raising=False)
    monkeypatch.delenv("COLLEGEHOUSE_SQL_PASSWORD", raising=False)


JAN = datetime.datetime(2026, 1, 1)
FEB = datetime.datetime(2026, 2, 1)


# ── _safe_ratio ──────────────────────────────────────────────────────────────

class TestSafeRatio:
    def test_basic(self):
        assert _safe_ratio(90, 100) == pytest.approx(0.9)

    def test_zero_denominator(self):
        assert _safe_ratio(90, 0) is None

    def test_none_inputs(self):
        assert _safe_ratio(None, 100) is None
        assert _safe_ratio(90, None) is None

    def test_non_numeric(self):
        assert _safe_ratio("abc", 100) is None


# ── fetch_market_performance ─────────────────────────────────────────────────

class TestFetch:
    def test_no_filters_returns_empty(self, creds, fake_pyodbc):
        assert fetch_market_performance() == []

    def test_missing_credentials_returns_empty(self, no_creds, fake_pyodbc):
        assert fetch_market_performance(institution="UCF") == []

    def test_pull_with_derived_metrics(self, creds, fake_pyodbc):
        fake_pyodbc._rows = [_raw_row(JAN)]
        rows = fetch_market_performance(institution="University of Central Florida")
        assert len(rows) == 1
        row = rows[0]
        assert row["BuildingName"] == "Hub Orlando"
        assert row["PreleasePct"] == pytest.approx(0.90)
        assert row["OccupancyPct"] == pytest.approx(0.95)
        assert row["RatePerBed"] == pytest.approx(900.0)
        assert row["RatePerSF"] == pytest.approx(3.0)
        assert fake_pyodbc._connections[-1].closed

    def test_filters_parameterized(self, creds, fake_pyodbc):
        fake_pyodbc._rows = []
        fetch_market_performance(
            institution="UCF", ipeds=132903, property_like=["Hub", "Verve"]
        )
        cursor = fake_pyodbc._connections[-1]._cursor
        assert "InstitutionName = ?" in cursor.executed
        assert "IPEDS = ?" in cursor.executed
        assert cursor.executed.count("BuildingName LIKE ?") == 2
        assert cursor.params == ["UCF", 132903, "%Hub%", "%Verve%"]

    def test_retries_then_succeeds(self, creds, fake_pyodbc, monkeypatch):
        monkeypatch.setattr(college_house.time, "sleep", lambda s: None)
        fake_pyodbc._fail_times = 1
        fake_pyodbc._rows = [_raw_row(JAN)]
        rows = fetch_market_performance(institution="UCF")
        assert len(rows) == 1

    def test_connection_failure_returns_empty(self, creds, fake_pyodbc, monkeypatch):
        monkeypatch.setattr(college_house.time, "sleep", lambda s: None)
        fake_pyodbc._fail_times = 99
        assert fetch_market_performance(institution="UCF") == []


# ── _resolve_driver ──────────────────────────────────────────────────────────

class TestResolveDriver:
    def _mod(self, drivers):
        mod = types.ModuleType("pyodbc")
        mod.drivers = lambda: drivers
        return mod

    def test_preferred_installed(self):
        mod = self._mod(["ODBC Driver 17 for SQL Server"])
        assert college_house._resolve_driver(
            "ODBC Driver 17 for SQL Server", mod
        ) == "ODBC Driver 17 for SQL Server"

    def test_falls_back_to_installed(self):
        mod = self._mod(["SQL Server", "ODBC Driver 18 for SQL Server"])
        assert college_house._resolve_driver(
            "ODBC Driver 17 for SQL Server", mod
        ) == "ODBC Driver 18 for SQL Server"

    def test_nothing_installed_keeps_preferred(self):
        mod = self._mod([])
        assert college_house._resolve_driver(
            "ODBC Driver 17 for SQL Server", mod
        ) == "ODBC Driver 17 for SQL Server"


# ── is_configured ────────────────────────────────────────────────────────────

class TestIsConfigured:
    def test_true_with_creds_and_pyodbc(self, creds, fake_pyodbc):
        assert is_configured() is True

    def test_false_without_creds(self, no_creds, fake_pyodbc):
        assert is_configured() is False


# ── summarize_latest_month ───────────────────────────────────────────────────

class TestSummarize:
    def test_latest_month_wins(self, creds, fake_pyodbc):
        fake_pyodbc._rows = [
            _raw_row(JAN, prelease=(50, 100)),
            _raw_row(FEB, prelease=(80, 100)),
        ]
        rows = fetch_market_performance(institution="UCF")
        summary = summarize_latest_month(rows)
        assert len(summary) == 1
        assert summary[0]["MonthDate"] == FEB
        assert summary[0]["PreleasePct"] == pytest.approx(0.80)

    def test_bed_weighted_across_bedroom_types(self, creds, fake_pyodbc):
        fake_pyodbc._rows = [
            _raw_row(FEB, bedrooms=2, beds=100, rate=(1000, 1)),
            _raw_row(FEB, bedrooms=4, beds=300, rate=(800, 1)),
        ]
        rows = fetch_market_performance(institution="UCF")
        summary = summarize_latest_month(rows)
        assert summary[0]["TotalBeds"] == 400
        # (1000*100 + 800*300) / 400 = 850
        assert summary[0]["RatePerBed"] == pytest.approx(850.0)

    def test_multiple_properties_sorted(self, creds, fake_pyodbc):
        fake_pyodbc._rows = [
            _raw_row(FEB, building="Verve Orlando"),
            _raw_row(FEB, building="Hub Orlando"),
        ]
        rows = fetch_market_performance(institution="UCF")
        summary = summarize_latest_month(rows)
        assert [s["BuildingName"] for s in summary] == ["Hub Orlando", "Verve Orlando"]


# ── compute_leasing_cycle_rent_growth ────────────────────────────────────────

class TestRentGrowth:
    """Rent growth uses LEASING-CYCLE AVERAGE rents: Sep → latest month vs
    the same Sep-to-month window one year prior."""

    def _rows(self, fake_pyodbc, raw):
        fake_pyodbc._rows = raw
        return fetch_market_performance(institution="UCF")

    def test_cycle_average_growth(self, creds, fake_pyodbc):
        rows = self._rows(fake_pyodbc, [
            # prior cycle window (Sep 2024 → Feb 2025): avg 950
            _raw_row(datetime.datetime(2024, 9, 1), rate=(900, 1)),
            _raw_row(datetime.datetime(2025, 2, 1), rate=(1000, 1)),
            # current cycle window (Sep 2025 → Feb 2026): avg 1000
            _raw_row(datetime.datetime(2025, 9, 1), rate=(980, 1)),
            _raw_row(datetime.datetime(2026, 2, 1), rate=(1020, 1)),
        ])
        growth = compute_leasing_cycle_rent_growth(rows)["Hub Orlando"]
        assert growth["current_avg_rent"] == pytest.approx(1000.0)
        assert growth["prior_avg_rent"] == pytest.approx(950.0)
        assert growth["rent_growth"] == pytest.approx(1000 / 950 - 1)

    def test_prior_window_clipped_to_same_months(self, creds, fake_pyodbc):
        """A prior-cycle month AFTER the current cutoff month must not count:
        latest = Feb 2026 → prior window is Sep 2024–Feb 2025, not Mar 2025."""
        rows = self._rows(fake_pyodbc, [
            _raw_row(datetime.datetime(2025, 2, 1), rate=(950, 1)),
            _raw_row(datetime.datetime(2025, 3, 1), rate=(5000, 1)),  # out of window
            _raw_row(datetime.datetime(2026, 2, 1), rate=(1000, 1)),
        ])
        growth = compute_leasing_cycle_rent_growth(rows)["Hub Orlando"]
        assert growth["prior_avg_rent"] == pytest.approx(950.0)

    def test_no_prior_data_returns_none(self, creds, fake_pyodbc):
        rows = self._rows(fake_pyodbc, [_raw_row(FEB)])
        growth = compute_leasing_cycle_rent_growth(rows)["Hub Orlando"]
        assert growth["rent_growth"] is None

    def test_growth_in_summary_and_outputs(self, creds, fake_pyodbc, tmp_path):
        rows = self._rows(fake_pyodbc, [
            _raw_row(datetime.datetime(2025, 2, 1), rate=(1000, 1)),
            _raw_row(datetime.datetime(2026, 2, 1), rate=(1050, 1)),
        ])
        summary = summarize_latest_month(rows)
        assert summary[0]["RentGrowthYoY"] == pytest.approx(0.05)

        text = format_market_performance_text(rows)
        assert "YoY Rent Growth" in text
        assert "LEASING-CYCLE AVERAGE" in text
        assert "5.0%" in text

        path = str(tmp_path / "extract.xlsx")
        write_extract_workbook(rows, path)
        ws = openpyxl.load_workbook(path)["Comp Performance Summary"]
        assert "YoY Rent Growth" in str(ws.cell(row=1, column=9).value)
        assert ws.cell(row=2, column=9).value == pytest.approx(0.05)


# ── format_market_performance_text ───────────────────────────────────────────

class TestFormatText:
    def test_empty_rows(self):
        assert format_market_performance_text([]) == ""

    def test_sections_and_formatting(self, creds, fake_pyodbc):
        fake_pyodbc._rows = [_raw_row(FEB)]
        rows = fetch_market_performance(institution="UCF")
        text = format_market_performance_text(rows)
        assert "MARKET DATA (from College House SQL" in text
        assert "TAB: Comp Performance Summary" in text
        assert "TAB: Monthly Performance By Property And Bedroom" in text
        assert "90.0%" in text      # prelease
        assert "95.0%" in text      # occupancy
        assert "$900" in text       # rate per bed
        assert "Hub Orlando" in text
        assert "2026-02" in text


# ── write_extract_workbook ───────────────────────────────────────────────────

class TestExtractWorkbook:
    def test_nothing_to_write(self, tmp_path):
        assert write_extract_workbook([], str(tmp_path / "x.xlsx")) is None

    def test_sheets_and_content(self, creds, fake_pyodbc, tmp_path):
        fake_pyodbc._rows = [_raw_row(FEB)]
        rows = fetch_market_performance(institution="UCF")
        path = str(tmp_path / "college_house_extract.xlsx")
        assert write_extract_workbook(rows, path) == path

        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Comp Performance Summary", "Monthly Raw Data"]

        summary = wb["Comp Performance Summary"]
        assert summary.cell(row=1, column=1).value == "Property"
        assert summary.cell(row=2, column=1).value == "Hub Orlando"
        assert summary.cell(row=2, column=5).value == pytest.approx(0.90)

        raw = wb["Monthly Raw Data"]
        header = [c.value for c in raw[1]]
        assert header[: len(MONTHLY_COLUMNS)] == MONTHLY_COLUMNS
        assert "PreleasePct" in header


# ── extract_college_house_market_data ────────────────────────────────────────

class TestExtractEntryPoint:
    def test_no_config_returns_empty(self):
        assert extract_college_house_market_data({}) == ""
        assert extract_college_house_market_data({"market_data": {}}) == ""

    def test_full_flow_writes_extract(self, creds, fake_pyodbc, tmp_path):
        fake_pyodbc._rows = [_raw_row(FEB)]
        cfg = {"market_data": {"college_house": {"institution": "UCF"}}}
        text = extract_college_house_market_data(cfg, output_dir=str(tmp_path))
        assert "MARKET DATA (from College House SQL" in text
        assert (tmp_path / "college_house_extract.xlsx").exists()

    def test_pull_failure_returns_empty(self, creds, fake_pyodbc, monkeypatch):
        monkeypatch.setattr(college_house.time, "sleep", lambda s: None)
        fake_pyodbc._fail_times = 99
        cfg = {"market_data": {"college_house": {"institution": "UCF"}}}
        assert extract_college_house_market_data(cfg) == ""
