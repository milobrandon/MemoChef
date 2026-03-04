"""Tests for market_workbook.py — workbook generation and data loading."""

import os
import tempfile

import pytest
from openpyxl import load_workbook

from market_workbook import (
    MarketDataWorkbook,
    PropertyInfo,
    RentComp,
    SaleComp,
    SubmarketStats,
    MacroEntry,
    UnitMixEntry,
    WorkbookMetadata,
    generate_workbook,
    load_sample_data,
)


@pytest.fixture
def sample_data():
    """Minimal but complete market data for testing."""
    return MarketDataWorkbook(
        property=PropertyInfo(
            name="Test Property",
            address="123 Main St",
            city="Testville",
            state="TX",
            submarket="Downtown",
            msa="Test-Metro",
            property_type="Multifamily",
        ),
        rent_comps=[
            RentComp(
                name=f"Comp {i}",
                address=f"{i}00 Oak Ave",
                distance_mi=0.5 * i,
                units=200 + i * 10,
                year_built=2020 + i,
                occupancy_pct=0.93 + i * 0.01,
                unit_rents={"1BR": 1500 + i * 100, "2BR": 2000 + i * 100},
                concessions="None",
                as_of_date="2026-02-01",
                source="Test Source",
            )
            for i in range(1, 5)
        ],
        sale_comps=[
            SaleComp(
                name=f"Sale {i}",
                address=f"{i}00 Elm St",
                sale_date=f"2025-0{i}-15",
                sale_price=50_000_000 + i * 5_000_000,
                units=200 + i * 20,
                price_per_unit=250_000 + i * 10_000,
                cap_rate=0.045 + i * 0.005,
                source="Test Source",
            )
            for i in range(1, 4)
        ],
        submarket=SubmarketStats(
            name="Downtown Test",
            avg_rent=1800,
            vacancy_pct=0.06,
            absorption_units=500,
            pipeline_units=2000,
            rent_growth_yoy_pct=0.03,
            median_hh_income=70000,
            population=50000,
            as_of_date="2026-Q1",
            source="Test Source",
        ),
        macro=[
            MacroEntry(
                date=f"2025-{m:02d}",
                sofr_rate=0.04 - m * 0.001,
                treasury_10yr=0.038 - m * 0.001,
                cpi_yoy=0.025 - m * 0.001,
                unemployment=0.04 - m * 0.0005,
                source="FRED",
            )
            for m in range(1, 7)
        ],
        unit_mix=[
            UnitMixEntry("Studio", 30, 500, 1400, 2.80),
            UnitMixEntry("1BR", 100, 720, 1750, 2.43),
            UnitMixEntry("2BR", 80, 1050, 2300, 2.19),
        ],
        metadata=WorkbookMetadata(
            prepared_by="Test User",
            prepared_date="2026-03-04",
            property_name="Test Property",
            notes="Test notes",
        ),
    )


class TestGenerateWorkbook:
    def test_creates_valid_xlsx(self, sample_data, tmp_path):
        out = str(tmp_path / "test_output.xlsx")
        result = generate_workbook(sample_data, out)
        assert os.path.exists(result)
        assert result == out

    def test_has_expected_tabs(self, sample_data, tmp_path):
        out = str(tmp_path / "test_tabs.xlsx")
        generate_workbook(sample_data, out)
        wb = load_workbook(out)
        sheet_names = wb.sheetnames
        assert "Cover" in sheet_names
        assert "Rent Comps" in sheet_names
        assert "Sale Comps" in sheet_names
        assert "Submarket Overview" in sheet_names
        assert "Macro Indicators" in sheet_names
        assert "Unit Mix Summary" in sheet_names
        assert "Sources & Notes" in sheet_names

    def test_cover_has_property_name(self, sample_data, tmp_path):
        out = str(tmp_path / "test_cover.xlsx")
        generate_workbook(sample_data, out)
        wb = load_workbook(out)
        ws = wb["Cover"]
        assert ws["B3"].value == "Market Data Workbook"
        # Property name should be in the sheet
        values = [ws.cell(row=r, column=3).value for r in range(5, 15)]
        assert "Test Property" in values

    def test_rent_comps_row_count(self, sample_data, tmp_path):
        out = str(tmp_path / "test_rent.xlsx")
        generate_workbook(sample_data, out)
        wb = load_workbook(out)
        ws = wb["Rent Comps"]
        # Header + 4 comps
        assert ws.cell(row=1, column=1).value == "Property"
        assert ws.cell(row=2, column=1).value == "Comp 1"
        assert ws.cell(row=5, column=1).value == "Comp 4"

    def test_sale_comps_formatting(self, sample_data, tmp_path):
        out = str(tmp_path / "test_sale.xlsx")
        generate_workbook(sample_data, out)
        wb = load_workbook(out)
        ws = wb["Sale Comps"]
        assert ws.cell(row=1, column=1).value == "Property"
        assert ws.cell(row=2, column=1).value == "Sale 1"

    def test_charts_exist_in_rent_comps(self, sample_data, tmp_path):
        out = str(tmp_path / "test_charts.xlsx")
        generate_workbook(sample_data, out)
        wb = load_workbook(out)
        ws = wb["Rent Comps"]
        assert len(ws._charts) >= 2  # Bar + Scatter

    def test_charts_exist_in_macro(self, sample_data, tmp_path):
        out = str(tmp_path / "test_macro_charts.xlsx")
        generate_workbook(sample_data, out)
        wb = load_workbook(out)
        ws = wb["Macro Indicators"]
        assert len(ws._charts) >= 2  # Rate trends + CPI

    def test_unit_mix_totals_row(self, sample_data, tmp_path):
        out = str(tmp_path / "test_unitmix.xlsx")
        generate_workbook(sample_data, out)
        wb = load_workbook(out)
        ws = wb["Unit Mix Summary"]
        # Total row should be at row 5 (header + 3 types + total)
        assert ws.cell(row=5, column=1).value == "Total"
        assert ws.cell(row=5, column=2).value == 210  # 30 + 100 + 80

    def test_empty_sections_skipped(self, tmp_path):
        minimal = MarketDataWorkbook(
            property=PropertyInfo("Min", "1 St", "City", "ST", "Sub", "MSA", "MF"),
        )
        out = str(tmp_path / "test_minimal.xlsx")
        generate_workbook(minimal, out)
        wb = load_workbook(out)
        assert "Cover" in wb.sheetnames
        assert "Rent Comps" not in wb.sheetnames  # No data → no sheet


class TestLoadSampleData:
    def test_loads_sample_file(self):
        sample_path = os.path.join(os.path.dirname(__file__), "sample_market_data.json")
        if not os.path.exists(sample_path):
            pytest.skip("sample_market_data.json not found")
        data = load_sample_data(sample_path)
        assert data.property.name == "Meridian at Oakwood"
        assert len(data.rent_comps) == 6
        assert len(data.sale_comps) == 5
        assert data.submarket is not None
        assert len(data.macro) == 12
        assert len(data.unit_mix) == 4

    def test_roundtrip_generate(self, tmp_path):
        sample_path = os.path.join(os.path.dirname(__file__), "sample_market_data.json")
        if not os.path.exists(sample_path):
            pytest.skip("sample_market_data.json not found")
        data = load_sample_data(sample_path)
        out = str(tmp_path / "roundtrip.xlsx")
        generate_workbook(data, out)
        assert os.path.exists(out)
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 7
