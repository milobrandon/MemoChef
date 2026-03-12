"""Tests for market workbook chart extraction."""
import pytest
from pathlib import Path


def test_extract_workbook_tables_format(tmp_path):
    """Verify output matches proforma text format."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rent Growth"
    ws["A1"] = "Year"
    ws["B1"] = "Submarket"
    ws["A2"] = 2024
    ws["B2"] = 3.5
    wb.save(tmp_path / "test.xlsx")

    from memo_chef.chart_extraction import extract_workbook_tables
    result = extract_workbook_tables(str(tmp_path / "test.xlsx"))
    assert "TAB: Rent Growth" in result
    assert "Row 1:" in result
    assert "Year" in result


def test_extract_workbook_tables_specific_tabs(tmp_path):
    """Only extract specified tabs."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Tab A"
    ws1["A1"] = "data"
    ws2 = wb.create_sheet("Tab B")
    ws2["A1"] = "other"
    wb.save(tmp_path / "test.xlsx")

    from memo_chef.chart_extraction import extract_workbook_tables
    result = extract_workbook_tables(str(tmp_path / "test.xlsx"), tab_names=["Tab A"])
    assert "TAB: Tab A" in result
    assert "TAB: Tab B" not in result


def test_extract_workbook_tables_empty_tabs(tmp_path):
    """Empty tabs are skipped."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empty"
    wb.save(tmp_path / "test.xlsx")

    from memo_chef.chart_extraction import extract_workbook_tables
    result = extract_workbook_tables(str(tmp_path / "test.xlsx"))
    assert "TAB:" not in result or result.strip() == ""
