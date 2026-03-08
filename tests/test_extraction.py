"""Tests for supplemental data extraction."""
import pytest
from memo_chef.extraction import extract_supplemental


def test_extract_excel_returns_text(tmp_path):
    """Excel extraction returns tab-delimited text."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    ws["A2"] = "HHI"
    ws["B2"] = 58200
    path = tmp_path / "test.xlsx"
    wb.save(path)

    result = extract_supplemental(str(path), "excel")
    assert "HHI" in result
    assert "58200" in result


def test_extract_pdf_returns_text(tmp_path):
    """PDF extraction returns page text."""
    from unittest.mock import patch, MagicMock

    mock_page = MagicMock()
    mock_page.extract_text.return_value = "Student affluence data: HHI $62,500"
    mock_page.extract_tables.return_value = []

    mock_pdf = MagicMock()
    mock_pdf.pages = [mock_page]
    mock_pdf.__enter__ = lambda s: s
    mock_pdf.__exit__ = MagicMock(return_value=False)

    with patch("pdfplumber.open", return_value=mock_pdf):
        result = extract_supplemental("fake.pdf", "pdf")
    assert "Student affluence" in result
    assert "62,500" in result


def test_extract_url_returns_text():
    """URL extraction returns page text content."""
    from unittest.mock import patch, MagicMock

    mock_resp = MagicMock()
    mock_resp.status_code = 200
    mock_resp.text = "<html><body><p>Market occupancy is 94.2%</p></body></html>"

    with patch("requests.get", return_value=mock_resp):
        result = extract_supplemental("https://example.com/data", "url")
    assert "94.2%" in result


def test_extract_unknown_type_raises():
    """Unknown source type raises ValueError."""
    with pytest.raises(ValueError, match="Unsupported source_type"):
        extract_supplemental("file.xyz", "unknown")


def test_extract_csv_returns_text(tmp_path):
    """CSV extraction returns row text."""
    csv_path = tmp_path / "data.csv"
    csv_path.write_text("Metric,Value\nHHI,58200\nPop Growth,2.1%\n")
    result = extract_supplemental(str(csv_path), "csv")
    assert "HHI" in result
    assert "58200" in result
