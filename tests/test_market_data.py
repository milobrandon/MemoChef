"""Tests for market data extraction and chart update functionality."""

import os
import sys

import openpyxl
import pytest

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from memo_automator import (
    extract_market_data,
    load_config,
)

from memo_chef.models import (
    MarketChartUpdate, MarketNarrativeUpdate,
    MarketTableCellUpdate, MarketTableUpdate, MarketDataUpdateSet,
)


class TestMarketDataModels:
    def test_chart_update_validates(self):
        u = MarketChartUpdate(
            page=3,
            chart_name="Rent Growth",
            series=[{"name": "Market A", "new_values": [1200, 1350], "old_values": [1100, 1250]}],
            categories=["2023", "2024"],
            add_series=[{"name": "Market D", "values": [900, 950]}],
            remove_series=["Market C"],
            source="Rent Growth tab",
            reasoning="Semantic match",
            confidence="high",
        )
        assert u.type == "chart_update"
        assert u.page == 3
        assert u.confidence == "high"

    def test_narrative_update_validates(self):
        u = MarketNarrativeUpdate(
            page=7,
            old_text="Rents grew 5%",
            new_text="Rents grew 12%",
            source="Rent Growth tab",
            reasoning="Updated figures",
            confidence="high",
        )
        assert u.type == "narrative_update"

    def test_table_update_validates(self):
        u = MarketTableUpdate(
            page=3,
            slide_table="Market Summary",
            updates=[{"row": 2, "col": 1, "old_value": "94%", "new_value": "96%"}],
            source="Tables tab",
            reasoning="Occupancy updated",
            confidence="medium",
        )
        assert u.type == "table_update"
        assert u.updates[0].row == 2

    def test_update_set_validates_mixed(self):
        s = MarketDataUpdateSet(
            market_data_updates=[
                {"type": "chart_update", "page": 3, "series": [], "source": "x", "reasoning": "y", "confidence": "high"},
                {"type": "narrative_update", "page": 5, "old_text": "a", "new_text": "b", "source": "x", "reasoning": "y", "confidence": "high"},
            ],
            unmatched_memo_metrics=["Absorption chart p9"],
            unmatched_workbook_tabs=["Backend"],
            warnings=["Low confidence match on p3"],
        )
        assert len(s.market_data_updates) == 2
        assert s.warnings == ["Low confidence match on p3"]

    def test_empty_update_set(self):
        s = MarketDataUpdateSet()
        assert s.market_data_updates == []
        assert s.warnings == []

    def test_update_set_accessor_helpers(self):
        s = MarketDataUpdateSet(
            market_data_updates=[
                {"type": "chart_update", "page": 3, "series": [], "source": "x", "reasoning": "y", "confidence": "high"},
                {"type": "narrative_update", "page": 5, "old_text": "a", "new_text": "b", "source": "x", "reasoning": "y", "confidence": "high"},
                {"type": "table_update", "page": 7, "slide_table": "T", "updates": [], "source": "x", "reasoning": "y", "confidence": "high"},
            ]
        )
        assert len(s.chart_updates()) == 1
        assert s.chart_updates()[0].page == 3
        assert len(s.narrative_updates()) == 1
        assert len(s.table_updates()) == 1


@pytest.fixture
def default_cfg():
    config_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config.yaml"
    )
    return load_config(config_path)


_LEGACY_DASHBOARD_TABS = [
    "Rent Comp Market",
    "Occupancy Comparison By Year",
    "Rent Growth Comparison By Year",
    "Rent Comp Survey",
    "Supply Demand Pipeline",
]


@pytest.fixture
def market_data_workbook(tmp_path):
    """Create a minimal market data workbook with market-keyword tabs."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _LEGACY_DASHBOARD_TABS[0]
    ws["A1"] = "IPEDS"
    ws["B1"] = 134130
    ws["A2"] = "University Name"
    ws["B2"] = "UF"

    for tab_name in _LEGACY_DASHBOARD_TABS[1:]:
        ws2 = wb.create_sheet(tab_name)
        ws2["A1"] = f"Header for {tab_name}"
        ws2["B1"] = 123.45

    # Add a back-end tab (should be ignored - no market keywords)
    ws3 = wb.create_sheet("PROPERTIES")
    ws3["A1"] = "This should not be extracted"

    path = str(tmp_path / "test_market_data.xlsx")
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def empty_workbook(tmp_path):
    """Create a workbook with no dashboard tabs."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RandomSheet"
    ws["A1"] = "Not a dashboard"
    path = str(tmp_path / "empty_market.xlsx")
    wb.save(path)
    wb.close()
    return path


class TestExtractMarketData:
    def test_happy_path(self, market_data_workbook, default_cfg):
        result = extract_market_data(market_data_workbook, default_cfg)
        assert "MARKET DATA (from workbook)" in result
        assert f"TAB: {_LEGACY_DASHBOARD_TABS[0]}" in result
        assert f"TAB: {_LEGACY_DASHBOARD_TABS[3]}" in result
        assert "IPEDS" in result
        assert "134130" in result

    def test_ignores_backend_tabs(self, market_data_workbook, default_cfg):
        result = extract_market_data(market_data_workbook, default_cfg)
        assert "PROPERTIES" not in result.split("MARKET DATA")[1]  # Only check after header
        assert "This should not be extracted" not in result

    def test_all_dashboard_tabs_extracted(self, market_data_workbook, default_cfg):
        result = extract_market_data(market_data_workbook, default_cfg)
        for tab in _LEGACY_DASHBOARD_TABS:
            assert f"TAB: {tab}" in result

    def test_no_dashboard_tabs_returns_empty(self, empty_workbook, default_cfg):
        result = extract_market_data(empty_workbook, default_cfg)
        assert result == ""

    def test_invalid_file_returns_empty(self, tmp_path, default_cfg):
        bad_path = str(tmp_path / "not_a_real_file.xlsx")
        with open(bad_path, "w") as f:
            f.write("not an excel file")
        result = extract_market_data(bad_path, default_cfg)
        assert result == ""

    def test_missing_file_returns_empty(self, default_cfg):
        result = extract_market_data("/nonexistent/path.xlsx", default_cfg)
        assert result == ""

    def test_partial_tabs(self, tmp_path, default_cfg):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rent Comp Market"
        ws["A1"] = "Data"
        ws2 = wb.create_sheet("Occupancy Growth")
        ws2["A1"] = "Comps"
        ws3 = wb.create_sheet("BACKEND_RAW")
        ws3["A1"] = "Internal only"
        path = str(tmp_path / "partial.xlsx")
        wb.save(path)
        wb.close()

        result = extract_market_data(path, default_cfg)
        assert "TAB: Rent Comp Market" in result
        assert "TAB: Occupancy Growth" in result
        assert "TAB: BACKEND_RAW" not in result

    def test_output_format_rows(self, market_data_workbook, default_cfg):
        """Verify tab-delimited row format matches proforma pattern."""
        result = extract_market_data(market_data_workbook, default_cfg)
        assert "Row 1:\t" in result  # Tab-delimited format


class TestRealMarketDataFile:
    REAL_FILE = "C:/Users/BrandonZmuda/Desktop/Claude/g. Memo Automator/a. Sandbox/New Template Test.xlsx"

    @pytest.mark.skipif(
        not os.path.exists(REAL_FILE),
        reason="Real market data file not available",
    )
    def test_real_file_extraction(self, default_cfg):
        result = extract_market_data(self.REAL_FILE, default_cfg)
        assert len(result) > 1000
        assert "MARKET DATA (from workbook)" in result

    @pytest.mark.skipif(
        not os.path.exists(REAL_FILE),
        reason="Real market data file not available",
    )
    def test_real_file_prompt_size(self, default_cfg):
        """Verify market data doesn't blow up prompt size unreasonably."""
        result = extract_market_data(self.REAL_FILE, default_cfg)
        # Dashboard tabs should be well under 50K chars
        assert len(result) < 50000, f"Market data too large: {len(result)} chars"


class TestDynamicExtractMarketData:
    """Tests for the new keyword-scoring extractor."""

    def _make_workbook(self, sheets: dict) -> str:
        """Create a temp xlsx with given {tab_name: [[row], [row]]} data."""
        import tempfile
        wb = openpyxl.Workbook()
        first = True
        for name, rows in sheets.items():
            if first:
                ws = wb.active
                ws.title = name
                first = False
            else:
                ws = wb.create_sheet(name)
            for row in rows:
                ws.append(row)
        path = tempfile.mktemp(suffix=".xlsx")
        wb.save(path)
        return path

    def test_scores_rent_tab_above_threshold(self):
        path = self._make_workbook({
            "Rent Comparison": [["Market", "Rent", "Occupancy"], ["A", 1200, 0.95]],
            "Backend Raw": [["id", "code"], [1, "X"]],
        })
        cfg = {"market_data": {"max_rows_per_tab": 50, "max_cols_per_tab": 10, "keyword_threshold": 2, "include_all_tabs": False}}
        result = extract_market_data(path, cfg)
        assert "Rent Comparison" in result
        assert "Backend Raw" not in result
        os.remove(path)

    def test_include_all_tabs_bypasses_scoring(self):
        path = self._make_workbook({
            "XYZ": [["col1", "col2"], ["a", "b"]],
        })
        cfg = {"market_data": {"max_rows_per_tab": 50, "max_cols_per_tab": 10, "keyword_threshold": 2, "include_all_tabs": True}}
        result = extract_market_data(path, cfg)
        assert "XYZ" in result
        os.remove(path)

    def test_falls_back_to_proforma_config_if_no_market_data_section(self):
        path = self._make_workbook({
            "Comp Set": [["Name", "Rent"], ["A", 1200]],
        })
        cfg = {"proforma": {"max_rows_per_tab": 50, "max_cols_per_tab": 10},
               "market_data": {"max_rows_per_tab": 50, "max_cols_per_tab": 10, "keyword_threshold": 1, "include_all_tabs": False}}
        result = extract_market_data(path, cfg)
        assert "Comp Set" in result
        os.remove(path)

    def test_missing_file_returns_empty(self):
        cfg = {"market_data": {"max_rows_per_tab": 50, "max_cols_per_tab": 10, "keyword_threshold": 2, "include_all_tabs": False}}
        assert extract_market_data("/no/such/file.xlsx", cfg) == ""

    def test_tab_header_line_present(self):
        path = self._make_workbook({
            "Occupancy Trend": [["Year", "Occupancy"], [2023, 0.94], [2024, 0.96]],
        })
        cfg = {"market_data": {"max_rows_per_tab": 50, "max_cols_per_tab": 10, "keyword_threshold": 1, "include_all_tabs": False}}
        result = extract_market_data(path, cfg)
        assert "TAB: Occupancy Trend" in result
        os.remove(path)
