"""Tests for the P3 proforma variant — only 'Presentation*' tabs are extracted."""

import os
import sys

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from memo_automator import (
    _is_p3_proforma,
    _select_p3_presentation_tabs,
    extract_proforma_data,
)


def _cfg():
    return {
        "proforma": {
            "tabs": ["Executive Summary", "Cash Flow"],
            "max_rows_per_tab": 250,
            "max_cols_per_tab": 20,
        }
    }


def _build_p3_workbook(path):
    """Synthesize a P3-shaped workbook: presentation tabs plus internal tabs."""
    wb = openpyxl.Workbook()
    # First sheet is internal — should not be extracted.
    internal = wb.active
    internal.title = "Model Cash Flow"
    internal["A1"] = "Internal Only"
    internal["B1"] = "DO_NOT_EXTRACT"

    pres_exec = wb.create_sheet("Presentation Exec Summary")
    pres_exec["A1"] = "Metric"
    pres_exec["B1"] = "Value"
    pres_exec["A2"] = "IRR"
    pres_exec["B2"] = "12.4%"

    pres_cf = wb.create_sheet("Presentation Cash Flow")
    pres_cf["A1"] = "Year"
    pres_cf["B1"] = "NOI"
    pres_cf["A2"] = 2027
    pres_cf["B2"] = 2_100_000

    # Tab whose name does not start with "Presentation" — should be excluded
    # even on a P3 workbook.
    other = wb.create_sheet("Pres Cash Flow RollUp")
    other["A1"] = "Should not appear"

    wb.save(path)
    wb.close()


class TestP3Detection:
    def test_filename_with_p3_prefix_detected(self):
        assert _is_p3_proforma(r"C:\foo\P3 Proforma_Orlando.xlsm") is True

    def test_filename_lowercase_p3_prefix_detected(self):
        assert _is_p3_proforma("/tmp/p3 proforma_deal.xlsm") is True

    def test_vanilla_proforma_not_detected(self):
        assert _is_p3_proforma(r"C:\foo\Proforma_Lexington.xlsm") is False

    def test_p3_substring_in_middle_not_detected(self):
        # Just the bare letters "p3" elsewhere in the name should not trigger.
        assert _is_p3_proforma(r"C:\foo\Proforma_with_p3_inside.xlsm") is False

    def test_p3_without_trailing_space_not_detected(self):
        # The convention is "P3 " (with a space). "P3-..." or "P3_..." aren't P3.
        assert _is_p3_proforma(r"C:\foo\P3-Foo.xlsm") is False
        assert _is_p3_proforma(r"C:\foo\P3_Bar.xlsm") is False


class TestP3TabSelection:
    def test_selects_presentation_prefixed_tabs(self):
        sheets = [
            "Executive Summary",
            "Model Cash Flow",
            "Presentation Exec Summary",
            "Presentation Cash Flow",
            "Presentation Dev Budget",
            "Pres Cash Flow RollUp",  # abbreviation — not "Presentation"
            "Assumptions",
        ]
        selected = _select_p3_presentation_tabs(sheets)
        assert selected == [
            "Presentation Exec Summary",
            "Presentation Cash Flow",
            "Presentation Dev Budget",
        ]

    def test_match_is_case_insensitive(self):
        sheets = ["presentation exec", "PRESENTATION CF", "pReSeNtAtIoN dev"]
        assert _select_p3_presentation_tabs(sheets) == sheets

    def test_no_presentation_tabs_returns_empty(self):
        assert _select_p3_presentation_tabs(["Sheet1", "Sheet2"]) == []


class TestP3Extraction:
    def test_p3_extraction_only_presentation_tabs(self, tmp_path):
        path = tmp_path / "P3 Proforma_Test.xlsm"
        _build_p3_workbook(str(path))

        text = extract_proforma_data(str(path), _cfg())

        # Both presentation tabs are extracted.
        assert "TAB: Presentation Exec Summary" in text
        assert "TAB: Presentation Cash Flow" in text
        assert "12.4%" in text
        assert "2100000" in text

        # Internal tabs are NOT extracted, even if the workbook contains them.
        assert "DO_NOT_EXTRACT" not in text
        assert "TAB: Model Cash Flow" not in text
        assert "TAB: Pres Cash Flow RollUp" not in text

    def test_p3_extraction_ignores_configured_tabs(self, tmp_path):
        """The cfg lists 'Executive Summary' / 'Cash Flow' (which exist in the
        workbook only as internal-style names). P3 detection must override the
        config to use Presentation* tabs instead."""
        path = tmp_path / "P3 Proforma_Override.xlsm"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Executive Summary"
        ws["A1"] = "INTERNAL_EXEC"
        wb.create_sheet("Cash Flow")["A1"] = "INTERNAL_CF"
        wb.create_sheet("Presentation Exec Summary")["A1"] = "PRES_EXEC_VAL"
        wb.save(path)
        wb.close()

        text = extract_proforma_data(str(path), _cfg())

        assert "PRES_EXEC_VAL" in text
        assert "INTERNAL_EXEC" not in text
        assert "INTERNAL_CF" not in text

    def test_p3_with_no_presentation_tabs_raises(self, tmp_path):
        path = tmp_path / "P3 Proforma_Empty.xlsm"
        wb = openpyxl.Workbook()
        wb.active.title = "Executive Summary"
        wb["Executive Summary"]["A1"] = "data"
        wb.save(path)
        wb.close()

        with pytest.raises(ValueError, match="no 'Presentation\\*' tabs"):
            extract_proforma_data(str(path), _cfg())

    def test_non_p3_proforma_uses_configured_tabs(self, tmp_path):
        """Vanilla proformas must still respect the configured tab list and
        ignore any incidental 'Presentation*' tabs."""
        path = tmp_path / "Proforma_Standard.xlsm"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Executive Summary"
        ws["A1"] = "Metric"
        ws["A2"] = "STANDARD_VALUE"
        cf = wb.create_sheet("Cash Flow")
        cf["A1"] = "CF_DATA"
        # Incidental presentation tab — must be ignored on a non-P3 workbook.
        pres = wb.create_sheet("Presentation Extra")
        pres["A1"] = "SHOULD_NOT_APPEAR"
        wb.save(path)
        wb.close()

        text = extract_proforma_data(str(path), _cfg())

        assert "STANDARD_VALUE" in text
        assert "CF_DATA" in text
        assert "SHOULD_NOT_APPEAR" not in text
        assert "TAB: Presentation Extra" not in text
